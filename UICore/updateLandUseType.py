import sys
import time
import traceback

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Side, Border, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from osgeo import ogr, gdal

from UICore import filegdbapi
from UICore.common import style_range
import UICore.fgdberror as fgdberror

from UICore.DataFactory import workspaceFactory
from UICore.Gv import DataType
from UICore.log4p import Log
import os
import copy

log = Log(__name__)

need_indexes = ["DLBM_index", "GHFLDM_index", "XZQDM_index", "GHFLSDL_index", "GHJGFLDM_index"]
region_dict = {
    '440303': '罗湖区',
    '440304': '福田区',
    '440305': '南山区',
    '440306': '宝安区',
    '440307': '龙岗区',
    '440308': '盐田区',
    '440309': '龙华区',
    '440310': '坪山区',
    '440311': '光明区',
    '440312': '大鹏新区'
}

report_dict = ["各区现状分类面积汇总表", "规划分类面积汇总表", "规划分类三大类面积汇总表", "规划结构分类面积汇总表"]

header_font = Font(bold=True, size=11)
header_font2 = Font(bold=True, size=9)
header_font3 = Font(bold=False, size=11)
cell_font = Font(bold=False, size=9)
cell_font2 = Font(bold=False, size=11)

border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                     bottom=Side(style='thin'))
alignment_center = Alignment(horizontal="center", vertical="center", wrapText=True)
alignment_right = Alignment(horizontal="right", vertical="center", wrapText=True)
alignment_right2 = Alignment(horizontal="right", vertical="center", wrapText=False)

# 大表头样式
header_style = NamedStyle(name="header_style")
# header_style.border = border_thin
header_style.alignment = alignment_center
header_style.font = header_font

# 小表头样式
header_style2 = NamedStyle(name="header_style2")
header_style2.border = border_thin
header_style2.alignment = alignment_center
header_style2.font = header_font2

# 灰色背景表头
header_style3 = NamedStyle(name="header_style3")
header_style3.border = border_thin
header_style3.alignment = alignment_center
header_style3.font = header_font
header_style3.fill = PatternFill("solid", fgColor="D9D9D9")

# 居右表头
header_style4 = NamedStyle(name="header_style4")
header_style4.alignment = alignment_right2
header_style4.font = header_font3

# 数据单元格居中样式
cell_center_style = NamedStyle(name="cell_center_style")
cell_center_style.border = border_thin
cell_center_style.alignment = alignment_center
cell_center_style.font = cell_font
cell_center_style.number_format = "0.00"

# 数据单元格居右样式
cell_right_style = NamedStyle(name="cell_right_style")
cell_right_style.border = border_thin
cell_right_style.alignment = alignment_right
cell_right_style.font = cell_font
cell_right_style.number_format = "0.00"

# 数据单元格居右加粗样式
cell_right_thick_style = NamedStyle(name="cell_right_thick_style")
cell_right_thick_style.border = border_thin
cell_right_thick_style.alignment = alignment_right
cell_right_thick_style.font = header_font2
cell_right_thick_style.number_format = "0.00"

# 报表2，3，4的非表头文字单元格格式
cell_common_style = NamedStyle(name="cell_common_style")
cell_common_style.border = border_thin
cell_common_style.alignment = alignment_center
cell_common_style.font = cell_font2

# 报表2，3，4的数字单元格格式
cell_number_style = NamedStyle(name="cell_number_style")
cell_number_style.border = border_thin
cell_number_style.alignment = alignment_center
cell_number_style.font = cell_font2
cell_number_style.number_format = "0.00"

step_log_color = "#0000FF"  # 右侧Logviewer中info日志的突出显示为蓝色

def update_and_stat(file_type, in_path, layer_name, right_header, rel_tables, MC_tables, DLBM_values, report_file_name,
                    bConvert, bReport1, bReport2, bReport3, bReport4):
    dataSource = None
    wks = None
    bflag = False

    cur_path, filename = os.path.split(os.path.abspath(sys.argv[0]))
    temp_sqliteDB_name = '%s.db' % (layer_name + '_' + time.strftime('%Y-%m-%d-%H-%M-%S'))
    temp_sqliteDB_path = os.path.join(cur_path, "tmp")

    bflag = True
    try:
        if bConvert:
            log.info("开始现状与国土空间规划分类转换...", color=step_log_color)
            start = time.time()

            if file_type == DataType.shapefile:
                bflag = update_attribute_value(file_type, in_path, layer_name, right_header, rel_tables)

            elif file_type == DataType.fileGDB:
                # bflag = update_attribute_value(file_type, in_path, layer_name, right_header, rel_tables)
                drop_index(in_path, layer_name, need_indexes)

                bflag = update_attribute_value_by_fileGDB(in_path, layer_name, right_header, rel_tables,
                                                          DLBM_values)

            end = time.time()
            log.info('现状与国土空间规划分类转换完成, 总共耗时:{}秒.'.format("{:.2f}".format(end - start)), color=step_log_color)

        if not bReport1 and not bReport2 and not bReport3 and not bReport4:
            return

        if not bflag:
            return

        # # 测试用
        # wks = workspaceFactory().get_factory(DataType.fileGDB)
        # dataSource = wks.openFromFile(in_path, 1)
        # layer = dataSource.GetLayerByName(layer_name)

        # log.info("开始进行报表统计和输出...", color=step_color)

        if file_type == DataType.fileGDB:
            wks = workspaceFactory().get_factory(DataType.fileGDB)
        elif file_type == DataType.shapefile:
            wks = workspaceFactory().get_factory(DataType.shapefile)
        dataSource = wks.openFromFile(in_path, 0)

        log.info("必要性字段检查...")
        bReports, bflag = check_report_need_fields([bReport1, bReport2, bReport3, bReport4], dataSource, layer_name)

        if not bflag:
            return

        log.info("创建用于统计的临时数据库...")
        dataSource = create_temp_sqliteDB(temp_sqliteDB_path, temp_sqliteDB_name, in_path, layer_name)

        if dataSource is not None:
            wb = Workbook()

            nReport = 0
            for bReport in bReports:
                if nReport == 0 and bReport:
                    start = time.time()
                    log.info('开始统计"各区现状分类面积汇总表"...', color=step_log_color)
                    bflag = output_stat_report1(wb, dataSource, layer_name, MC_tables)
                    if not bflag:
                        return

                    end = time.time()
                    log.info('"{}"统计完成, 总共耗时:{}秒.'.format(report_dict[nReport], "{:.2f}".format(end - start)), color=step_log_color)
                    wb.save(report_file_name)

                if nReport == 1 and bReport:
                    start = time.time()
                    log.info('开始统计"规划分类面积汇总表"...', color=step_log_color)
                    # if file_type == DataType.fileGDB:
                    #     drop_index(in_path, layer_name, need_indexes)
                    bflag = output_stat_report2(wb, dataSource, layer_name)
                    if not bflag:
                        return

                    end = time.time()
                    wb.save(report_file_name)
                    log.info('"{}"统计完成, 总共耗时:{}秒.'.format(report_dict[nReport], "{:.2f}".format(end - start)), color=step_log_color)

                if nReport == 2 and bReport:
                    start = time.time()
                    log.info('开始统计"规划分类三大类面积汇总表"...', color=step_log_color)
                    # if file_type == DataType.fileGDB:
                    #     drop_index(in_path, layer_name, need_indexes)
                    bflag = output_stat_report3(wb, dataSource, layer_name)
                    if not bflag:
                        return
                    end = time.time()
                    log.info('"{}"统计完成, 总共耗时:{}秒.'.format(report_dict[nReport], "{:.2f}".format(end - start)), color=step_log_color)
                    wb.save(report_file_name)

                if nReport == 3 and bReport:
                    start = time.time()
                    log.info('开始统计"规划结构分类面积汇总表"...', color=step_log_color)
                    # if file_type == DataType.fileGDB:
                    #     drop_index(in_path, layer_name, need_indexes)
                    bflag = output_stat_report4(wb, dataSource, layer_name)
                    if not bflag:
                        return
                    end = time.time()
                    log.info('"{}"统计完成, 总共耗时:{}秒.'.format(report_dict[nReport], "{:.2f}".format(end - start)), color=step_log_color)
                    wb.save(report_file_name)

                nReport += 1

            log.info("所有报表都已统计完成，结果保存至路径{}".format(report_file_name), color=step_log_color)
    except:
        log.error(traceback.format_exc())
        return
    finally:
        del wks
        del dataSource
        if file_type == DataType.fileGDB:
            drop_index(in_path, layer_name, need_indexes)
        # elif file_type == DataType.shapefile:
        remove_temp_sqliteDB(temp_sqliteDB_path, temp_sqliteDB_name)


# 创建用于统计的临时数据库
def create_temp_sqliteDB(temp_sqliteDB_path, temp_sqliteDB_name, in_path, layer_name):
    dataSource = None
    if not os.path.exists(temp_sqliteDB_path):
        os.mkdir(temp_sqliteDB_path)
    translateOptions = gdal.VectorTranslateOptions(format="SQLite", layerName=layer_name,
                                                   datasetCreationOptions=["SPATIALITE=NO"])
    hr = gdal.VectorTranslate(os.path.join(temp_sqliteDB_path, temp_sqliteDB_name), in_path,
                              options=translateOptions)
    if not hr:
        raise Exception("创建临时数据库出错!错误原因:\n{}".format(traceback.format_exc()))
    else:
        wks = workspaceFactory().get_factory(DataType.sqlite)
        dataSource = wks.openFromFile(os.path.join(temp_sqliteDB_path, temp_sqliteDB_name), 1)
    del hr

    return dataSource


def remove_temp_sqliteDB(in_path, db_name):
    db_path = os.path.join(in_path, db_name)
    try:
        if os.path.exists(db_path):
            os.remove(db_path)
    except:
        log.warning("删除临时sqlite数据库文件{}出错，可能是数据库文件被占用，请手动删除!".format(db_path))


def update_attribute_value(file_type, in_path, layer_name, right_header, rel_tables, DLBM_values=None):
    # layer = dataSource.GetLayer(0)
    layer = None
    dataSource = None
    wks = None

    try:
        # start = time.time()

        if file_type == DataType.shapefile:
            wks = workspaceFactory().get_factory(DataType.shapefile)
        elif file_type == DataType.fileGDB:
            wks = workspaceFactory().get_factory(DataType.fileGDB)

        # wks = workspaceFactory().get_factory(DataType.shapefile)
        dataSource = wks.openFromFile(in_path, 1)
        layer = dataSource.GetLayer(0)

        layerDefn = layer.GetLayerDefn()

        field_names = []
        for i in range(layerDefn.GetFieldCount()):
            fieldName = layerDefn.GetFieldDefn(i).GetName()
            field_names.append(fieldName)

        right_header.append("XZQDM")
        right_header.append("XZQMC")

        log.info("第1步: 根据规则表的DLBM右侧表头增加矢量图层{}中的相应字段...".format(layer_name))
        for header_value in right_header:
            if header_value not in field_names:
                new_field = ogr.FieldDefn(header_value, ogr.OFTString)
                new_field.SetWidth(200)
                layer.CreateField(new_field, True)
                del new_field

        iprop = 1
        total_count = layer.GetFeatureCount()

        feature = layer.GetNextFeature()
        field_index_order = []
        for field_name in right_header:
            field_index = feature.GetFieldIndex(field_name)
            # field_dict[field_name] = field_index
            field_index_order.append(field_index)

        XZQDM_index = feature.GetFieldIndex("XZQDM")
        XZQMC_index = feature.GetFieldIndex("XZQMC")
        ZLDWDM_index = feature.GetFieldIndex("ZLDWDM")
        ZLDWMC_index = feature.GetFieldIndex("ZLDWMC")

        log.info("第2步: 根据规则表更新{}图层对应数据...".format(layer_name))
        icount = 0

        lack_BM = set()
        while feature:
            DLBM_value = feature.GetField("DLBM")
            bchecked = False

            # ZLDWDM = feature.GetField("ZLDWDM")
            # ZLDWMC = feature.GetField("ZLDWMC")
            ZLDWDM = feature.GetField(ZLDWDM_index)
            ZLDWMC = feature.GetField(ZLDWMC_index)

            if ZLDWDM is not None and len(ZLDWDM) > 6:
                XZQDM = ZLDWDM[0:6]

                if XZQDM in region_dict:
                    XZQMC = region_dict[XZQDM]
                else:
                    XZQMC = ""

                if XZQDM == '440307':
                    if ZLDWMC != '宝龙街道' and ZLDWMC != '布吉街道' and ZLDWMC != '龙城街道' and ZLDWMC != '龙岗街道' \
                            and ZLDWMC != '平湖街道' and ZLDWMC != '坪地街道' and ZLDWMC != '园山街道' and ZLDWMC != '南湾街道' \
                            and ZLDWMC != '坂田街道' and ZLDWMC != '吉华街道' and ZLDWMC != '横岗街道':
                        # feature.SetField("XZQDM", "440312")
                        # feature.SetField("XZQMC", "大鹏新区")
                        feature.SetField(XZQDM_index, "440312")
                        feature.SetField(XZQMC_index, "大鹏新区")
                    else:
                        feature.SetField(XZQDM_index, XZQDM)
                        feature.SetField(XZQMC_index, XZQMC)
                else:
                    feature.SetField(XZQDM_index, XZQDM)
                    feature.SetField(XZQMC_index, XZQMC)

            for i in range(len(rel_tables)):
                rel = rel_tables[i]
                field_name = right_header[i]
                field_index = field_index_order[i]

                if DLBM_value not in rel:
                    if not bchecked:
                        if DLBM_value not in lack_BM:
                            lack_BM.add(DLBM_value)
                            log.warning("出现了在规则表中不存在的编码：{}".format(DLBM_value))
                        # log.warning("第{}个要素的地类编码{}在规则表中不存在！".format(icount, DLBM_value))
                        feature.SetField(field_index, None)
                        bchecked = True
                    else:
                        feature.SetField(field_index, None)
                else:
                    if feature.GetFieldType(field_index) == ogr.OFTString:
                        feature.SetField(field_index, str(rel[DLBM_value]))
                    elif feature.GetFieldType(field_index) == ogr.OFTInteger:
                        feature.SetField(field_index, int(rel[DLBM_value]))
                    elif feature.GetFieldType(field_index) == ogr.OFTReal:
                        feature.SetField(field_index, float(rel[DLBM_value]))
                    else:
                        log.error("第{}个要素的字段{}是无法识别的数据类型. 字段类型只允许是整型、字符型或者浮点型，请调整原始数据!".format(icount, field_name))
                        feature.SetField(field_index, None)

            # # 如果是CZCSXM是201或者202则重新赋值
            # CZCSXM_index = feature.GetFieldIndex("CZCSXM")
            # if CZCSXM_index > 0:
            #     CZCSXM_value = feature.GetField(CZCSXM_index)
            #     if str(CZCSXM_value).strip() == '201' or str(CZCSXM_value).strip() == '202':
            #         feature.SetField("XZFLSDL", "农用地/未利用地")
            #         feature.SetField("GHJGFLDM", "07")
            #         feature.SetField("GHJGFLMC", "城乡建设用地")
            #         feature.SetField("SFJSYD", "是")

            layer.SetFeature(feature)
            feature = layer.GetNextFeature()

            icount += 1
            if int(icount * 100 / total_count) == iprop * 20:
                log.info("{:.0%}已处理完成...".format(icount / total_count))
                iprop += 1

        # end = time.time()
        # log.info("操作完成, 总共耗时:{}秒.".format("{:.2f}".format(end-start)))
        return True
    except:
        log.error("无法更新数据！错误原因:\n{}".format(traceback.format_exc()))
        return False
        # return False, "无法更新数据！错误原因:\n{}".format(traceback.format_exc())
    finally:
        del dataSource
        del layer
        del feature
        del wks


# 采用执行SQL方式更新数据
def update_attribute_value_by_fileGDB(in_path, layer_name, right_header, rel_tables, DLBM_values):
    layer = None
    dataSource = None
    feature = None
    wks = None
    exec_res = None

    need_add_header = copy.deepcopy(right_header)

    try:
        # start = time.time()
        wks = workspaceFactory().get_factory(DataType.fileGDB)
        dataSource = wks.openFromFile(in_path, 1)
        layer = dataSource.GetLayerByName(layer_name)

        layerDefn = layer.GetLayerDefn()

        field_names = []
        for i in range(layerDefn.GetFieldCount()):
            fieldName = layerDefn.GetFieldDefn(i).GetName()
            field_names.append(fieldName)

        #  增加两列表示行政区代码和行政区名称
        # if "XZQDM" not in field_names:
        #     new_field = ogr.FieldDefn("XZQDM", ogr.OFTString)
        #     new_field.SetWidth(100)
        #     layer.CreateField(new_field, True)
        #     del new_field
        #
        # if "XZQMC" not in field_names:
        #     new_field = ogr.FieldDefn("XZQMC", ogr.OFTString)
        #     new_field.SetWidth(100)
        #     layer.CreateField(new_field, True)
        #     del new_field
        need_add_header.append("XZQMC")
        need_add_header.append("XZQDM")

        log.info("第1步: 根据规则表的DLBM右侧表头增加矢量图层{}中的相应字段...".format(layer_name))
        for header_value in need_add_header:
            if header_value not in field_names:
                new_field = ogr.FieldDefn(header_value, ogr.OFTString)
                new_field.SetWidth(100)
                layer.CreateField(new_field, True)
                del new_field

        log.info("第2步: 对矢量图层{}的字段创建索引...".format(layer_name))
        exec_str = r"CREATE INDEX DLBM_index ON {} (DLBM)".format(layer_name)
        exec_res = dataSource.ExecuteSQL(exec_str)
        dataSource.ReleaseResultSet(exec_res)

        log.info("第3步: 计算矢量图层{}中DLBM字段的唯一值...".format(layer_name))
        exec_str = r"SELECT DISTINCT DLBM FROM {}".format(layer_name)
        exec_res = dataSource.ExecuteSQL(exec_str, dialect="SQLite")

        DLBM_keys = []
        feature = exec_res.GetNextFeature()
        while feature:
            DLBM_key = feature.GetField(0)
            DLBM_keys.append(DLBM_key)
            feature = exec_res.GetNextFeature()

        feature = layer.GetFeature(1)
        dataSource.ReleaseResultSet(exec_res)

        # exec_str = r"UPDATE {} SET ZLDWDM_1=''".format(layer_name)
        # exec_res = dataSource.ExecuteSQL(exec_str)
        # dataSource.ReleaseResultSet(exec_res)
        log.info("第4步: 更新行政区代码和行政区名称...")

        exec_str = r"UPDATE {} SET XZQDM=SUBSTRING(ZLDWDM, 1, 6)".format(layer_name)
        exec_res = dataSource.ExecuteSQL(exec_str)
        dataSource.ReleaseResultSet(exec_res)

        exec_str = r"UPDATE {} SET XZQDM='440312' WHERE ZLDWDM LIKE '440307%' AND " \
                   r"ZLDWMC <> '宝龙街道' AND ZLDWMC <> '布吉街道' AND ZLDWMC <> '龙城街道' AND " \
                   r"ZLDWMC <> '龙岗街道' AND ZLDWMC <> '平湖街道' AND ZLDWMC <> '坪地街道' AND " \
                   r"ZLDWMC <> '园山街道' AND ZLDWMC <> '南湾街道' AND ZLDWMC <> '坂田街道' AND " \
                   r"ZLDWMC <> '吉华街道' AND ZLDWMC <> '横岗街道'".format(layer_name)
        exec_res = dataSource.ExecuteSQL(exec_str)
        dataSource.ReleaseResultSet(exec_res)

        exec_str = r"CREATE INDEX XZQDM_index ON {} (XZQDM)".format(layer_name)
        exec_res = dataSource.ExecuteSQL(exec_str)
        dataSource.ReleaseResultSet(exec_res)

        for key, value in region_dict.items():
            exec_str = r"UPDATE {} SET XZQMC='{}' WHERE XZQDM='{}'".format(layer_name, value, key)
            # print(exec_str)
            exec_res = dataSource.ExecuteSQL(exec_str)
            dataSource.ReleaseResultSet(exec_res)

        log.info("第5步: 根据规则表计算矢量图层相应字段的值...")

        for DLBM_key in DLBM_keys:
            log.info("更新矢量图层{}字段DLBM中所有等于{}的值".format(layer_name, DLBM_key))

            if DLBM_key not in DLBM_values:
                log.warning("字段DLBM中包含规则表中不存在的编码{}".format(DLBM_key))

            for i in range(len(rel_tables)):
                rel = rel_tables[i]
                field_name = right_header[i]

                if DLBM_key in rel:
                    if feature.GetFieldType(field_name) == ogr.OFTString:
                        exec_str = r"UPDATE {} SET {} = '{}' WHERE DLBM = '{}'".format(layer_name, field_name,
                                                                                       rel[DLBM_key], DLBM_key)
                    elif feature.GetFieldType(field_name) == ogr.OFTInteger or feature.GetFieldType(
                            field_name) == ogr.OFTReal:
                        exec_str = r"UPDATE {} SET {} = {} WHERE DLBM = '{}'".format(layer_name, field_name,
                                                                                     rel[DLBM_key], DLBM_key)
                    else:
                        log.error("图层{}的字段{}是无法识别的数据类型. 字段类型只允许是整型、字符型或者浮点型，请调整原始数据!".format(layer_name, field_name))
                        exec_str = r"UPDATE {} SET {} = NULL WHERE DLBM = '{}'".format(layer_name, field_name, DLBM_key)
                else:
                    exec_str = r"UPDATE {} SET {} = NULL WHERE DLBM = '{}'".format(layer_name, field_name, DLBM_key)

                exec_res = dataSource.ExecuteSQL(exec_str)
                dataSource.ReleaseResultSet(exec_res)

        # log.info("第5步: 特别更新字段CZCSXM等于201或202的对应字段值...".format(layer_name, DLBM_key))
        # CZCSXM_index = feature.GetFieldIndex("CZCSXM")
        # if CZCSXM_index > 0:
        #     XZFLSDL_index = feature.GetFieldIndex("XZFLSDL")
        #     GHJGFLDM_index = feature.GetFieldIndex("GHJGFLDM")
        #     GHJGFLMC_index = feature.GetFieldIndex("GHJGFLMC")
        #     SFJSYD_index = feature.GetFieldIndex("SFJSYD")
        #
        #     exec_str = r"CREATE INDEX CZCSXM_index ON {} (CZCSXM)".format(layer_name)
        #     exec_res = dataSource.ExecuteSQL(exec_str)
        #     dataSource.ReleaseResultSet(exec_res)
        #
        #     if XZFLSDL_index > -1:
        #         exec_str = r"UPDATE {} SET XZFLSDL = '{}' WHERE CZCSXM = '201' or CZCSXM = '202'".format(layer_name,
        #                                                                                                  '农用地/未利用地')
        #         exec_res = dataSource.ExecuteSQL(exec_str)
        #         dataSource.ReleaseResultSet(exec_res)
        #     if GHJGFLDM_index > -1:
        #         exec_str = r"UPDATE {} SET GHJGFLDM = '{}' WHERE CZCSXM = '201' or CZCSXM = '202'".format(layer_name,
        #                                                                                                   '07')
        #         exec_res = dataSource.ExecuteSQL(exec_str)
        #         dataSource.ReleaseResultSet(exec_res)
        #     if GHJGFLMC_index > -1:
        #         exec_str = r"UPDATE {} SET GHJGFLMC = '{}' WHERE CZCSXM = '201' or CZCSXM = '202'".format(layer_name,
        #                                                                                                   '城乡建设用地')
        #         exec_res = dataSource.ExecuteSQL(exec_str)
        #         dataSource.ReleaseResultSet(exec_res)
        #     if SFJSYD_index > -1:
        #         exec_str = r"UPDATE {} SET SFJSYD = '{}' WHERE CZCSXM = '201' or CZCSXM = '202'".format(layer_name, '是')
        #         exec_res = dataSource.ExecuteSQL(exec_str)
        #         dataSource.ReleaseResultSet(exec_res)

        # end = time.time()
        # log.info("操作完成, 总共耗时:{}秒".format("{:.2f}".format(end-start)))
        return True
    except:
        # return False, "无法更新数据！错误原因:\n{}".format(traceback.format_exc())
        log.error("无法更新数据！错误原因:\n{}".format(traceback.format_exc()))
        return False
    finally:
        del dataSource
        del layer
        del feature
        del wks


def drop_index(in_path, layer_name, indexes):
    gdb = filegdbapi.Geodatabase()
    hr = filegdbapi.OpenGeodatabase(in_path, gdb)
    if hr != fgdberror.S_OK:
        raise Exception("读取数据库{}错误，错误代码:{}".format(in_path, str(hr)))
    tlb = filegdbapi.Table()
    hr = gdb.OpenTable(layer_name, tlb)
    if hr != fgdberror.S_OK:
        raise Exception("打开图层{}错误，错误代码:{}".format(layer_name, str(hr)))

    for index in indexes:
        tlb.DeleteIndex(index)

    gdb.CloseTable(tlb)
    filegdbapi.CloseGeodatabase(gdb)


def check_report_need_fields(bReports, dataSource, layer_name):
    layer = dataSource.GetLayerByName(layer_name)

    bflag = False
    res = copy.deepcopy(bReports)
    report_need_fields = []

    nReport = 0
    for bReport in bReports:
        if bReport:
            if nReport == 0:
                report_need_fields = ['DLMC', 'DLBM', 'XZFLSDL', 'ZLDWDM', 'XZQDM', 'ZLDWMC', 'TBDLMJ', 'KCMJ']
            elif nReport == 1:
                report_need_fields = ['TBDLMJ', 'KCMJ', 'GHFLDM1', 'GHFLMC1', 'GHFLDM2', 'GHFLMC2']
            elif nReport == 2:
                report_need_fields = ['TBDLMJ', 'GHFLSDL']
            elif nReport == 3:
                report_need_fields = ['GHJGFLDM', 'TBDLMJ']

            all_field_names = check_field(layer, report_need_fields, nReport)
            if all_field_names is None:
                res[nReport] = False
            else:
                res[nReport] = True
                bflag = True
        else:
            res[nReport] = False

        nReport += 1

    del layer
    return res, bflag

# 报表1 各区现状面积汇总表
def output_stat_report1(wb, dataSource, layer_name, MC_tables):
    # report_need_fields = ['DLMC', 'XZFLSDL', 'GHFLDM1', 'GHFLMC1', 'GHFLDM2', 'GHFLMC2', 'GHFLSDL', 'GHJGFLDM', 'GHJGFLMC', 'SFJSYD', 'ZLDWDM_1', 'ZLDWMC', 'TBMJ', 'KCMJ']

    layer_name = "[{}]".format(layer_name)

    # if file_type == DataType.shapefile:
    try:
        # check_report_need_fields(1, dataSource, layer_name)

        ws = wb.create_sheet('表2 各区现状分类面积汇总表')

        region_names = ['深圳市', '罗湖区', '福田区', '南山区', '宝安区', '龙岗区', '盐田区', '龙华区', '坪山区', '光明区', '大鹏新区']
        region_codes = ['4403', '440303', '440304', '440305', '440306', '440307', '440308', '440309', '440310',
                        '440311', '440312']
        col_count = len(region_names) + 2

        log.info("第1步：创建表头...")
        # 注意： 要先设置样式再合并，否则边框会出问题，这是openpyxl的Bug， 相关讨论见https://foss.heptapod.net/openpyxl/openpyxl/-/issues/365
        ws.cell(1, 1).value = "表2 各区现状分类面积汇总表"
        ws.cell(1, 1).style = header_style
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)

        ws.cell(2, col_count).value = "单位：公顷"
        ws.cell(2, col_count).style = header_style4

        ws.cell(3, 1).value = "行政区域"
        ws.cell(3, 1).style = header_style2
        ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
        #
        ws.cell(3, 2).value = "名称"
        ws.cell(3, 2).style = header_style2
        ws.cell(4, 2).value = "代码"
        ws.cell(4, 2).style = header_style2

        ws.cell(5, 1).value = "国土调查总面积"
        ws.cell(5, 1).style = header_style2
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=2)

        ws.cell(6, 1).value = "三大类"
        ws.cell(6, 1).style = header_style2
        ws.merge_cells(start_row=6, start_column=1, end_row=8, end_column=1)

        ws.cell(6, 2).value = "农用地"
        ws.cell(6, 2).style = cell_center_style
        ws.cell(7, 2).value = "建设用地"
        ws.cell(7, 2).style = cell_center_style
        ws.cell(8, 2).value = "未利用地"
        ws.cell(8, 2).style = cell_center_style

        for i in range(3, col_count + 1):
            ws.cell(3, i).value = region_names[i - 3]
            ws.cell(3, i).style = header_style2

            ws.cell(4, i).value = region_codes[i - 3]
            ws.cell(4, i).style = cell_center_style

            ws.cell(5, i).style = header_style2

            # exec_str = r"SELECT SUM(TBMJ) FROM {} WHERE XZFLSDL='{}' AND SUBSTR(ZLDWDM, 1, 4)='{}'".format(layer_name, ws.cell(6, 2).value, region_codes[i - 3])
            # exec_layer = dataSource.ExecuteSQL(exec_str, dialect="SQLite")
            # exec_res = exec_layer.GetNextFeature().GetField(0)
            # ws.cell(6, i).value = exec_res

        log.info("第2步：统计各区现状三大类面积...")

        # 统计三大类面积
        for i in range(0, 3):
            exec_str = r"SELECT SUBSTR(XZQDM, 1, 6), SUM(TBDLMJ) FROM {} WHERE XZFLSDL='{}' GROUP BY SUBSTR(XZQDM, 1, 6)".format(
                layer_name, str(ws.cell(6 + i, 2).value).strip())
            # print(exec_str)
            exec_layer = dataSource.ExecuteSQL(exec_str, dialect="SQLite")

            if exec_layer is None:
                log.warning("{}执行结果为空!".format(exec_str))
            else:
                exec_res = exec_layer.GetNextFeature()

                while exec_res:
                    ZDDWDM = exec_res.GetField(0)
                    ZDDWDM_MJ = '%.2f' % (exec_res.GetField(1) / 10000)

                    pos = region_codes.index(ZDDWDM)
                    if pos > -1:
                        ws.cell(6 + i, pos + 3).value = float(ZDDWDM_MJ)
                        ws.cell(6 + i, pos + 3).style = cell_right_style
                    else:
                        log.warning("没有相应的区域代码{}!".format(ZDDWDM))

                    exec_res = exec_layer.GetNextFeature()

                dataSource.ReleaseResultSet(exec_layer)
                del exec_res

        log.info("第3步：统计各区分类面积...")
        i = 0
        start_row = 9
        xiaoji_rows = []
        for mc, DLMCs in MC_tables.items():
            ws.cell(start_row, 1).value = "{}\n({})".format(mc, str(i).zfill(2))
            ws.cell(start_row, 1).style = header_style2

            ws.cell(start_row, 2).value = "小计\n({})".format(str(i).zfill(2))
            ws.cell(start_row, 2).style = header_style2
            xiaoji_rows.append(start_row)

            for j in range(len(DLMCs)):
                ws.cell(start_row + j + 1, 2).value = "{}\n({})".format(DLMCs[j]["DLMC"], DLMCs[j]["DLBM"])
                ws.cell(start_row + j + 1, 2).style = cell_center_style

                for iRegion in range(3, col_count + 1):
                    ws.cell(start_row + j + 1, iRegion).value = float('%.2f' % 0.00)
                    ws.cell(start_row + j + 1, iRegion).style = cell_right_style

                DLBM_key = DLMCs[j]["DLBM"]
                if DLBM_key == '1203':  # 田坎的面积单独计算
                    exec_str = r"SELECT SUBSTR(XZQDM, 1, 6), SUM(KCMJ) FROM {} GROUP BY SUBSTR(XZQDM, 1, 6)".format(
                        layer_name, DLBM_key)
                else:
                    exec_str = r"SELECT SUBSTR(XZQDM, 1, 6), SUM(TBDLMJ) FROM {} WHERE DLBM='{}' GROUP BY SUBSTR(XZQDM, 1, 6)".format(
                        layer_name, DLBM_key)
                exec_layer = dataSource.ExecuteSQL(exec_str, dialect="SQLite")

                if exec_layer is None:
                    log.warning("{}执行结果为空!".format(exec_str))
                else:
                    exec_res = exec_layer.GetNextFeature()

                    while exec_res:
                        ZDDWDM = exec_res.GetField(0)
                        ZDDWDM_MJ = '%.2f' % (exec_res.GetField(1) / 10000)

                        pos = region_codes.index(ZDDWDM)
                        if pos > -1:
                            ws.cell(start_row + j + 1, pos + 3).value = float(ZDDWDM_MJ)
                            if DLBM_key == '1203':  # 把田坎面积加入到三大类的农用地面积中
                                if ws.cell(6, pos + 3).value == "":
                                    ws.cell(6, pos + 3).value = float(ZDDWDM_MJ)
                                else:
                                    ws.cell(6, pos + 3).value = ws.cell(6, pos + 3).value + float(ZDDWDM_MJ)
                        else:
                            log.warning("没有相应的区域代码{}!".format(ZDDWDM))
                        ws.cell(start_row + j + 1, pos + 3).style = cell_right_style

                        exec_res = exec_layer.GetNextFeature()

                    dataSource.ReleaseResultSet(exec_layer)
                    del exec_res

            # 二级分类小计
            for iRegion in range(3, col_count + 1):
                sum = 0
                for j in range(len(DLMCs)):
                    sum = sum + float(ws.cell(start_row + j + 1, iRegion).value)
                ws.cell(start_row, iRegion).value = float('%.2f' % sum)
                ws.cell(start_row, iRegion).style = cell_right_thick_style

            if len(DLMCs) > 1:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + len(DLMCs), end_column=1)
                start_row = start_row + len(DLMCs) + 1  # 增加一行"小计"
            else:
                start_row = start_row + len(DLMCs)
            i += 1

        # 深圳市总计
        log.info("第4步：汇总各区统计结果...")
        for i in range(6, start_row):
            sum = 0
            for iRange in range(4, col_count + 1):
                if ws.cell(i, iRange).value is not None:
                    sum = sum + float(ws.cell(i, iRange).value)
            ws.cell(i, 3).value = float('%.2f' % sum)
            if i in xiaoji_rows:
                ws.cell(i, 3).style = cell_right_thick_style
            else:
                ws.cell(i, 3).style = cell_right_style

        for iRange in range(3, col_count + 1):  # 国土调查总面积
            sum = 0
            for i in range(0, 3):
                if ws.cell(6 + i, iRange).value is not None:
                    sum = sum + float(ws.cell(6 + i, iRange).value)
            ws.cell(5, iRange).value = float(sum)
            ws.cell(5, iRange).style = cell_right_thick_style

        return True
    except:
        log.error("无法完成报表统计！错误原因:\n{}".format(traceback.format_exc()))
        return False
        # return False, "无法完成报表统计！错误原因:\n{}".format(traceback.format_exc())


# 报表2 规划分类面积汇总表 格式写死
def output_stat_report2(wb, dataSource, layer_name):
    report_need_fields = ['TBDLMJ', 'KCMJ', 'GHFLDM1', 'GHFLMC1', 'GHFLDM2', 'GHFLMC2']

    col_count = 5

    try:
        # layer = dataSource.GetLayerByName(layer_name)

        # log.info("第1步：必要性字段检查...")
        # all_field_names = check_field(file_type, dataSource, layer, report_need_fields, report=2)

        # if all_field_names is None:
        #     return

        ws = wb.create_sheet('表3 规划分类面积汇总表')

        # if file_type == DataType.shapefile:
        layer_name = "[{}]".format(layer_name)

        log.info("第1步：创建表头...")
        # 注意： 要先设置样式再合并，否则边框会出问题，这是openpyxl的Bug， 相关讨论见https://foss.heptapod.net/openpyxl/openpyxl/-/issues/365
        ws.cell(1, 1).value = "表3 各区现状分类面积汇总表"
        ws.cell(1, 1).style = header_style
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)

        ws.cell(2, col_count).value = "单位：公顷"
        ws.cell(2, col_count).style = header_style4

        ws.cell(3, 1).value = "GHFLDM1"
        ws.cell(3, 2).value = "GHFLMC1"
        ws.cell(3, 3).value = "GHFLDM2"
        ws.cell(3, 4).value = "GHFLMC2"
        ws.cell(3, 5).value = "GHFLMJ"
        for i in range(1, 6):
            ws.cell(3, i).style = header_style3

        ws.cell(4, 1).value = "01"
        ws.cell(4, 2).value = "耕地"
        ws.cell(5, 1).value = "02"
        ws.cell(5, 2).value = "园地"
        ws.cell(6, 1).value = "03"
        ws.cell(6, 2).value = "林地"
        ws.cell(7, 1).value = "04"
        ws.cell(7, 2).value = "草地"
        ws.cell(8, 1).value = "05"
        ws.cell(8, 2).value = "湿地"
        ws.cell(9, 1).value = "06"
        ws.cell(9, 2).value = "农业设施建设用地"
        for i in range(4, 10):
            for j in range(1, 5):
                ws.cell(i, j).style = cell_common_style

        ws.cell(10, 1).value = "07"
        ws.cell(10, 1).style = cell_common_style
        ws.merge_cells(start_row=10, end_row=12, start_column=1, end_column=1)
        ws.cell(10, 2).value = "居住用地"
        ws.cell(10, 2).style = cell_common_style
        ws.merge_cells(start_row=10, end_row=12, start_column=2, end_column=2)
        ws.cell(10, 3).value = "0701"
        ws.cell(10, 3).style = cell_common_style
        ws.cell(10, 4).value = "城镇住宅用地"
        ws.cell(10, 4).style = cell_common_style
        ws.cell(11, 3).value = "0703"
        ws.cell(11, 3).style = cell_common_style
        ws.cell(11, 4).value = "农村宅基地"
        ws.cell(11, 4).style = cell_common_style
        ws.cell(12, 3).value = "小计"
        ws.cell(12, 3).style = cell_common_style
        ws.merge_cells(start_row=12, end_row=12, start_column=3, end_column=4)

        ws.cell(13, 1).value = "08"
        ws.cell(13, 1).style = cell_common_style
        ws.cell(13, 2).value = "公共管理与公共服务用地"
        ws.cell(13, 2).style = cell_common_style
        ws.cell(13, 3).style = cell_common_style
        ws.cell(13, 4).style = cell_common_style

        ws.cell(14, 1).value = "09"
        ws.cell(14, 1).style = cell_common_style
        ws.cell(14, 2).value = "商业服务业用地"
        ws.cell(14, 2).style = cell_common_style
        ws.cell(14, 3).style = cell_common_style
        ws.cell(14, 4).style = cell_common_style

        ws.cell(15, 1).value = "10"
        ws.cell(15, 1).style = cell_common_style
        ws.merge_cells(start_row=15, end_row=17, start_column=1, end_column=1)
        ws.cell(15, 2).value = "工矿用地"
        ws.cell(15, 2).style = cell_common_style
        ws.merge_cells(start_row=15, end_row=17, start_column=2, end_column=2)
        ws.cell(15, 3).value = "1001"
        ws.cell(15, 3).style = cell_common_style
        ws.cell(15, 4).value = "工业用地"
        ws.cell(15, 4).style = cell_common_style
        ws.cell(16, 3).value = "1002"
        ws.cell(16, 3).style = cell_common_style
        ws.cell(16, 4).value = "采矿用地"
        ws.cell(16, 4).style = cell_common_style
        ws.cell(17, 3).value = "小计"
        ws.cell(17, 3).style = cell_common_style
        ws.merge_cells(start_row=17, end_row=17, start_column=3, end_column=4)

        ws.cell(18, 1).value = "11"
        ws.cell(18, 1).style = cell_common_style
        ws.cell(18, 2).value = "仓储用地"
        ws.cell(18, 2).style = cell_common_style
        ws.cell(18, 3).style = cell_common_style
        ws.cell(18, 4).style = cell_common_style

        ws.cell(19, 1).value = "12"
        ws.cell(19, 1).style = cell_common_style
        ws.merge_cells(start_row=19, end_row=27, start_column=1, end_column=1)
        ws.cell(19, 2).value = "交通运输用地"
        ws.cell(19, 2).style = cell_common_style
        ws.merge_cells(start_row=19, end_row=27, start_column=2, end_column=2)
        ws.cell(19, 3).value = "1201"
        ws.cell(19, 4).value = "铁路用地"
        ws.cell(20, 3).value = "1202"
        ws.cell(20, 4).value = "公路用地"
        ws.cell(20, 3).value = "1202"
        ws.cell(20, 4).value = "公路用地"
        ws.cell(21, 3).value = "1203"
        ws.cell(21, 4).value = "机场用地"
        ws.cell(22, 3).value = "1204"
        ws.cell(22, 4).value = "港口码头用地"
        ws.cell(23, 3).value = "1205"
        ws.cell(23, 4).value = "管道运输用地"
        ws.cell(24, 3).value = "1206"
        ws.cell(24, 4).value = "城市轨道交通用地"
        ws.cell(25, 3).value = "1207"
        ws.cell(25, 4).value = "城镇道路用地"
        ws.cell(26, 3).value = "1208"
        ws.cell(26, 4).value = "交通场站用地"
        for i in range(19, 27):
            for j in range(3, 5):
                ws.cell(i, j).style = cell_common_style
        ws.cell(27, 3).value = "小计"
        ws.cell(27, 3).style = cell_common_style
        ws.merge_cells(start_row=27, end_row=27, start_column=3, end_column=4)

        ws.cell(28, 1).value = "13"
        ws.cell(28, 1).style = cell_common_style
        ws.merge_cells(start_row=28, end_row=30, start_column=1, end_column=1)
        ws.cell(28, 2).value = "公共设施用地"
        ws.cell(28, 2).style = cell_common_style
        ws.merge_cells(start_row=28, end_row=30, start_column=2, end_column=2)
        ws.cell(28, 3).value = "1312"
        ws.cell(28, 3).style = cell_common_style
        ws.cell(28, 4).value = "水工设施用地"
        ws.cell(28, 4).style = cell_common_style
        ws.cell(29, 3).value = "-"
        ws.cell(29, 3).style = cell_common_style
        ws.cell(29, 4).value = "其他公用设施用地（不含水工设施用地）"
        ws.cell(29, 4).style = cell_common_style
        ws.cell(30, 3).value = "小计"
        ws.cell(30, 3).style = cell_common_style
        ws.merge_cells(start_row=30, end_row=30, start_column=3, end_column=4)

        ws.cell(31, 1).value = "14"
        ws.cell(31, 2).value = "绿地与开敞空间用地"
        ws.cell(32, 1).value = "15"
        ws.cell(32, 2).value = "特殊用地"
        ws.cell(33, 1).value = "16"
        ws.cell(33, 2).value = "留白用地"
        ws.cell(34, 1).value = "17"
        ws.cell(34, 2).value = "陆地水域"
        for i in range(31, 35):
            for j in range(1, 5):
                ws.cell(i, j).style = cell_common_style

        ws.cell(35, 1).value = "23"
        ws.cell(35, 1).style = cell_common_style
        ws.merge_cells(start_row=35, end_row=38, start_column=1, end_column=1)
        ws.cell(35, 2).value = "其他土地"
        ws.cell(35, 2).style = cell_common_style
        ws.merge_cells(start_row=35, end_row=38, start_column=2, end_column=2)
        ws.cell(35, 3).value = "2301"
        ws.cell(35, 3).style = cell_common_style
        ws.cell(35, 4).value = "空闲地"
        ws.cell(35, 4).style = cell_common_style
        ws.cell(36, 3).value = "2302"
        ws.cell(36, 3).style = cell_common_style
        ws.cell(36, 4).value = "田坎"
        ws.cell(36, 4).style = cell_common_style
        ws.cell(37, 3).value = "-"
        ws.cell(37, 3).style = cell_common_style
        ws.cell(37, 4).value = "其他（不含空闲地、田坎）"
        ws.cell(37, 4).style = cell_common_style
        ws.cell(38, 3).value = "小计"
        ws.cell(38, 3).style = cell_common_style
        ws.merge_cells(start_row=38, end_row=38, start_column=3, end_column=4)

        ws.cell(39, 1).value = "合计"
        ws.cell(39, 1).style = cell_common_style
        ws.merge_cells(start_row=39, end_row=39, start_column=1, end_column=4)

        for i in range(4, 40):
            ws.cell(i, 5).style = cell_number_style

        log.info("第2步: 对矢量图层{}的字段创建索引...".format(layer_name))
        exec_str = r"CREATE INDEX GHFLDM_index ON {} (GHFLDM1, GHFLDM2)".format(layer_name)
        exec_res = dataSource.ExecuteSQL(exec_str)
        dataSource.ReleaseResultSet(exec_res)
        del exec_res

        log.info("第3步：统计规划分类面积...")

        total_MJ = 0
        GHFLDM1_rows = [4, 5, 6, 7, 8, 9, 13, 14, 18, 31, 32, 33, 34]
        for i in GHFLDM1_rows:
            GHFLDM1 = str(ws.cell(i, 1).value).strip()
            exec_str = "SELECT SUM(TBDLMJ) FROM {} WHERE GHFLDM1='{}'".format(layer_name, GHFLDM1)
            MJ = stat_mj_by_sql(dataSource, exec_str)
            ws.cell(i, 5).value = MJ
            total_MJ = total_MJ + MJ

        GHFLDM2_rows = [10, 11, 15, 16, 19, 20, 21, 22, 23, 24, 25, 26, 28]
        for i in GHFLDM2_rows:
            GHFLDM2 = str(ws.cell(i, 3).value).strip()
            exec_str = "SELECT SUM(TBDLMJ) FROM {} WHERE GHFLDM2='{}'".format(layer_name, GHFLDM2)
            MJ = stat_mj_by_sql(dataSource, exec_str)
            ws.cell(i, 5).value = MJ

        ws.cell(12, 5).value = ws.cell(10, 5).value + ws.cell(11, 5).value
        ws.cell(17, 5).value = ws.cell(15, 5).value + ws.cell(16, 5).value
        sum = 0
        for i in range(19, 27):
            sum = sum + ws.cell(i, 5).value
        ws.cell(27, 5).value = sum

        exec_str = "SELECT SUM(TBDLMJ) FROM {} WHERE GHFLDM1='13' AND GHFLDM2<>'1312'".format(layer_name)
        MJ = stat_mj_by_sql(dataSource, exec_str)
        ws.cell(29, 5).value = MJ
        ws.cell(30, 5).value = ws.cell(28, 5).value + ws.cell(29, 5).value

        exec_str = "SELECT SUM(TBDLMJ) FROM {} WHERE GHFLDM2=='2301'".format(layer_name)
        MJ = stat_mj_by_sql(dataSource, exec_str)
        exec_str = "SELECT SUM(KCMJ) FROM {}".format(layer_name)
        KCMJ = stat_mj_by_sql(dataSource, exec_str)
        ws.cell(36, 5).value = KCMJ
        exec_str = "SELECT SUM(TBDLMJ) FROM {} WHERE GHFLDM1='23' AND GHFLDM2 <> '2301' AND GHFLDM2 <> '2302'".format(
            layer_name)
        MJ_23 = stat_mj_by_sql(dataSource, exec_str)
        ws.cell(35, 5).value = MJ
        ws.cell(37, 5).value = MJ_23
        ws.cell(38, 5).value = MJ_23 + MJ + KCMJ

        total_MJ = total_MJ + ws.cell(27, 5).value
        total_MJ = total_MJ + ws.cell(17, 5).value
        total_MJ = total_MJ + ws.cell(12, 5).value
        total_MJ = total_MJ + ws.cell(38, 5).value
        total_MJ = total_MJ + ws.cell(30, 5).value

        ws.cell(39, 5).value = total_MJ

        ws.column_dimensions[get_column_letter(1)].width = 15
        ws.column_dimensions[get_column_letter(2)].width = 20
        ws.column_dimensions[get_column_letter(3)].width = 15
        ws.column_dimensions[get_column_letter(4)].width = 20
        ws.column_dimensions[get_column_letter(5)].width = 25

        return True
    except:
        log.error("无法完成报表统计！错误原因:\n{}".format(traceback.format_exc()))
        return False


# 报表3 规划分类三大类面积汇总表
def output_stat_report3(wb, dataSource, layer_name):
    # report_need_fields = ['TBDLMJ', 'GHFLSDL']
    col_count = 2

    try:
        # layer = dataSource.GetLayerByName(layer_name)

        # log.info("第1步：必要性字段检查...")
        # all_field_names = check_field(file_type, dataSource, layer, report_need_fields, report=2)
        #
        # if all_field_names is None:
        #     return

        ws = wb.create_sheet('表4 规划分类三大类面积汇总表')

        # if file_type == DataType.shapefile:
        layer_name = "[{}]".format(layer_name)

        log.info("第1步：创建表头...")
        ws.cell(1, 1).value = "表4 规划分类三大类面积汇总表"
        ws.cell(1, 1).style = header_style
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)

        ws.cell(2, col_count).value = "单位：公顷"
        ws.cell(2, col_count).style = header_style4

        ws.cell(3, 1).value = "GHFLSDL"
        ws.cell(3, 2).value = "GHFLMJ"
        for i in range(1, 3):
            ws.cell(3, i).style = header_style3

        ws.cell(4, 1).value = "农用地"
        ws.cell(5, 1).value = "建设用地"
        ws.cell(6, 1).value = "未利用地"

        log.info("第2步: 对矢量图层{}的字段创建索引...".format(layer_name))
        exec_str = r"CREATE INDEX GHFLSDL_index ON {} (GHFLSDL)".format(layer_name)
        exec_res = dataSource.ExecuteSQL(exec_str)
        dataSource.ReleaseResultSet(exec_res)
        del exec_res

        log.info("第3步：统计规划分类三大类面积...")
        for i in range(4, 7):
            ws.cell(i, 1).style = cell_common_style
            GHFLSDL = str(ws.cell(i, 1).value).strip()
            exec_str = "SELECT SUM(TBDLMJ) FROM {} WHERE GHFLSDL=='{}'".format(layer_name, GHFLSDL)
            MJ = stat_mj_by_sql(dataSource, exec_str)
            ws.cell(i, 2).value = MJ
            ws.cell(i, 2).style = cell_common_style

        ws.column_dimensions[get_column_letter(1)].width = 15
        ws.column_dimensions[get_column_letter(2)].width = 25

        return True
    except:
        log.error("无法完成报表统计！错误原因:\n{}".format(traceback.format_exc()))
        return False


def output_stat_report4(wb, dataSource, layer_name):
    # report_need_fields = ['GHJGFLDM', 'TBDLMJ']
    col_count = 3

    try:
        # layer = dataSource.GetLayerByName(layer_name)

        # log.info("第1步：必要性字段检查...")
        # all_field_names = check_field(file_type, dataSource, layer, report_need_fields, report=2)
        #
        # if all_field_names is None:
        #     return

        ws = wb.create_sheet('表5 规划结构分类面积汇总表')

        # if file_type == DataType.shapefile:
        layer_name = "[{}]".format(layer_name)

        log.info("第1步：创建表头...")
        ws.cell(1, 1).value = "表5 规划结构分类面积汇总表"
        ws.cell(1, 1).style = header_style
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)

        ws.cell(2, col_count).value = "单位：公顷"
        ws.cell(2, col_count).style = header_style4

        ws.cell(3, 1).value = "GHJGFLDM"
        ws.cell(3, 2).value = "GHJGFLMC"
        ws.cell(3, 3).value = "GHJGFLMJ"
        for i in range(1, 4):
            ws.cell(3, i).style = header_style3

        ws.cell(4, 1).value = "01"
        ws.cell(5, 1).value = "02"
        ws.cell(6, 1).value = "03"
        ws.cell(7, 1).value = "04"
        ws.cell(8, 1).value = "05"
        ws.cell(9, 1).value = "06"
        ws.cell(10, 1).value = "07"
        ws.cell(11, 1).value = "08"
        ws.cell(12, 1).value = "09"
        ws.cell(13, 1).value = "15"
        ws.cell(14, 1).value = "16"

        ws.cell(4, 2).value = "耕地"
        ws.cell(5, 2).value = "园地"
        ws.cell(6, 2).value = "林地"
        ws.cell(7, 2).value = "草地"
        ws.cell(8, 2).value = "湿地"
        ws.cell(9, 2).value = "农业设施建设用地"
        ws.cell(10, 2).value = "城乡建设用地"
        ws.cell(11, 2).value = "区域基础设施用地"
        ws.cell(12, 2).value = "其他建设用地"
        ws.cell(13, 2).value = "陆地水域"
        ws.cell(14, 2).value = "其他土地"

        for i in range(4, 15):
            for j in range(1, 3):
                ws.cell(i, j).style = cell_common_style

        log.info("第2步: 对矢量图层{}的字段创建索引...".format(layer_name))
        exec_str = r"CREATE INDEX GHJGFLDM_index ON {} (GHJGFLDM)".format(layer_name)
        exec_res = dataSource.ExecuteSQL(exec_str)
        dataSource.ReleaseResultSet(exec_res)
        del exec_res

        log.info("第3步：统计规划结构分类面积...")
        for i in range(4, 15):
            ws.cell(i, 1).style = cell_common_style
            GHJGFLDM = str(ws.cell(i, 1).value).strip()
            exec_str = "SELECT SUM(TBDLMJ) FROM {} WHERE GHJGFLDM=='{}'".format(layer_name, GHJGFLDM)
            MJ = stat_mj_by_sql(dataSource, exec_str)
            ws.cell(i, 3).value = MJ
            ws.cell(i, 3).style = cell_common_style

        ws.column_dimensions[get_column_letter(1)].width = 20
        ws.column_dimensions[get_column_letter(2)].width = 20
        ws.column_dimensions[get_column_letter(3)].width = 20

        return True
    except:
        log.error("无法完成报表统计！错误原因:\n{}".format(traceback.format_exc()))
        return False


def stat_mj_by_sql(dataSource, exec_str):
    MJ = ""
    exec_layer = dataSource.ExecuteSQL(exec_str, dialect="SQLite")
    if exec_layer is None:
        log.warning("{}执行结果为空!".format(exec_str))
    else:
        exec_res = exec_layer.GetNextFeature()

        while exec_res:
            val = exec_res.GetField(0)
            if val is not None:
                MJ = '%.2f' % (val / 10000)
            else:
                MJ = '%.2f' % 0
            break
    dataSource.ReleaseResultSet(exec_layer)
    del exec_layer

    if MJ == "":
        return 0
    else:
        return float(MJ)


# def output_stat_report2(file_type, wb, dataSource, layer_name, rel_tables, right_header):
#     report_need_fields = ['TBMJ', 'GHFLDM1', 'GHFLMC1', 'GHFLDM2', 'GHFLMC2']
#
#     col_count = 5
#
#     try:
#         layer = dataSource.GetLayerByName(layer_name)
#
#         log.info("第1步：必要性字段检查...")
#         all_field_names = check_field(file_type, dataSource, layer, report_need_fields, report=2)
#
#         if all_field_names is None:
#             return
#
#         ws = wb.create_sheet('规划分类面积汇总表')
#
#         if file_type == DataType.shapefile:
#             layer_name = "[{}]".format(layer_name)
#
#         GHFLDM1_index = -1
#         GHFLMC1_index = -1
#         GHFLDM2_index = -1
#         GHFLMC2_index = -1
#         for i in range(len(rel_tables)):
#             field_name = right_header[i]
#             if field_name.upper() == 'GHFLDM1':
#                 GHFLDM1_index = i
#             elif field_name.upper() == 'GHFLMC1':
#                 GHFLMC1_index = i
#             elif field_name.upper() == 'GHFLDM2':
#                 GHFLDM2_index = i
#             elif field_name.upper() == 'GHFLMC2':
#                 GHFLMC2_index = i
#
#         MC_tables = {}  # 存放一级规划分类代码（GHFLDM1）和GHFLMC1，GHFLDM2，GHFLMC2之间的关系表
#         # for i in range(len(rel_tables[GHFLDM1_index])):
#         GHFLDM1_keys = []
#         for DLBM, GHFLDM1 in rel_tables[GHFLDM1_index].items():
#             GHFLDM1_keys.append(GHFLDM1)
#
#             GHFLMC1 = rel_tables[GHFLMC1_index][DLBM]
#             GHFLDM2 = rel_tables[GHFLDM2_index][DLBM]
#             GHFLMC2 = rel_tables[GHFLMC2_index][DLBM]
#
#             if GHFLMC1 == '':
#                 continue
#
#             if GHFLDM1 not in MC_tables:
#                 dm2 = set()
#                 mc2 = set()
#                 dm2.add(GHFLDM2)
#                 mc2.add(GHFLMC2)
#                 MC_tables[GHFLDM1] = {
#                     "GHFLMC1": GHFLMC1,
#                     "GHFLDM2": dm2,
#                     "GHFLMC2": mc2
#                 }
#             else:
#                 MC_tables[GHFLDM1]["GHFLDM2"].add(GHFLDM2)
#                 MC_tables[GHFLDM1]["GHFLMC2"].add(GHFLMC2)
#                 MC_tables[GHFLDM1]["GHFLMC1"] = GHFLMC1
#
#
#         log.info("第2步：创建表头...")
#         # 注意： 要先设置样式再合并，否则边框会出问题，这是openpyxl的Bug， 相关讨论见https://foss.heptapod.net/openpyxl/openpyxl/-/issues/365
#         ws.cell(1, 1).value = "表2 各区现状分类面积汇总表"
#         ws.cell(1, 1).style = header_style
#         ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
#
#         ws.cell(2, col_count).value = "单位：公顷"
#         ws.cell(2, col_count).font = header_font3
#
#         ws.cell(3, 1).value = "GHFLDM1"
#         ws.cell(3, 2).value = "GHFLMC1"
#         ws.cell(3, 3).value = "GHFLDM2"
#         ws.cell(3, 4).value = "GHFLMC2"
#         ws.cell(3, 5).value = "GHFLMJ"
#         for i in range(1, 6):
#             ws.cell(3, i).style = header_style3
#
#         start_row = 4
#         xiaoji_rows = []  # 记录小计的行号
#         for GHFLDM1, val in MC_tables.items():
#             GHFLMC1 = val["GHFLMC1"]
#             GHFLDM2_lst = list(val["GHFLDM2"])
#             GHFLMC2_lst = list(val["GHFLMC2"])
#
#             GHFLDM1_len = len(GHFLDM2_lst)
#
#             ws.cell(start_row, 1).value = GHFLDM1
#             ws.cell(start_row, 2).value = GHFLMC1
#
#             for i in range(len(GHFLDM2_lst)):
#                 ws.cell(start_row + i, 3).value = GHFLDM2_lst[i]
#                 ws.cell(start_row + i, 4).value = GHFLMC2_lst[i]
#
#             if GHFLDM1_len > 1:
#                 ws.cell(start_row + GHFLDM1_len, 3). value = "小计"
#                 ws.merge_cells(start_row=start_row + GHFLDM1_len, end_row=start_row + GHFLDM1_len, start_column=3, end_column=4)
#
#                 # "公共设施用地"和"其他用地"特别对待，只保留两行
#                 if GHFLDM1 == '13':
#                     GHFLDM1_len = 2
#                 if GHFLDM1 == '23':
#                     GHFLDM1_len = 3
#
#                 ws.merge_cells(start_row=start_row, end_row=start_row + GHFLDM1_len, start_column=1, end_column=1)
#                 ws.merge_cells(start_row=start_row, end_row=start_row + GHFLDM1_len, start_column=2, end_column=2)
#
#                 start_row = start_row + GHFLDM1_len + 1
#             elif GHFLDM1_len == 1:
#                 start_row = start_row + GHFLDM1_len
#
#
#         print("OK")
#         return True
#     except:
#         log.error(traceback.format_exc())
#         return False

def check_field(layer, report_need_fields, nReport):
    all_field_names = []
    layer_name = layer.GetName()

    # layer_name = "[{}]".format(layer_name)

    berror = False
    layerDefn = layer.GetLayerDefn()

    for i in range(layerDefn.GetFieldCount()):
        fieldName = layerDefn.GetFieldDefn(i).GetName()
        # if fieldName.upper() == 'ZLDWDM':
        #     all_field_names.append('ZLDWDM_1')
        all_field_names.append(fieldName.upper())

    # if report == 1:
    #     if layerDefn.GetFieldIndex('ZLDWDM_1') < 0:
    #         new_field = ogr.FieldDefn('ZLDWDM_1', ogr.OFTString)
    #         new_field.SetWidth(200)
    #         layer.CreateField(new_field, True)
    #         del new_field

    for need_field in report_need_fields:
        if need_field not in all_field_names:
            log.warning('输出报表"{}"缺失必要字段"{}"，无法执行输出报表操作，请补全！'.format(report_dict[nReport], need_field))
            berror = True

    # if report == 1:
    #     # 这里有BUG，需要先给一个值，让新字段不为空，然后才能复制其他字段的值
    #     exec_str = r"UPDATE {} SET ZLDWDM_1=''".format(layer_name)
    #     exec_res = dataSource.ExecuteSQL(exec_str)
    #     dataSource.ReleaseResultSet(exec_res)
    #     exec_str = r"UPDATE {} SET ZLDWDM_1=ZLDWDM".format(layer_name)
    #     exec_res = dataSource.ExecuteSQL(exec_str)
    #     dataSource.ReleaseResultSet(exec_res)
    #     exec_str = r"UPDATE {} SET ZLDWDM_1='4403120000000000000' WHERE ZLDWDM LIKE '440307%' AND " \
    #                r"ZLDWMC <> '宝龙街道' AND ZLDWMC <> '布吉街道' AND ZLDWMC <> '龙城街道' AND " \
    #                r"ZLDWMC <> '龙岗街道' AND ZLDWMC <> '平湖街道' AND ZLDWMC <> '坪地街道' AND " \
    #                r"ZLDWMC <> '园山街道' AND ZLDWMC <> '南湾街道' AND ZLDWMC <> '坂田街道' AND " \
    #                r"ZLDWMC <> '吉华街道' AND ZLDWMC <> '横岗街道'".format(layer_name)
    #     exec_res = dataSource.ExecuteSQL(exec_str)
    #     dataSource.ReleaseResultSet(exec_res)

    if berror:
        return None
    else:
        return all_field_names
