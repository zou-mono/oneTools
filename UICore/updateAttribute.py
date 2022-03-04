import time
import traceback

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Side, Border, NamedStyle
from openpyxl.worksheet.cell_range import CellRange
from osgeo import ogr
from UICore.common import style_range

from UICore.DataFactory import workspaceFactory
from UICore.Gv import DataType
from UICore.log4p import Log

log = Log(__name__)


report_need_fields = ['DLMC', 'XZFLSDL', 'GHFLDM1', 'GHFLMC1', 'GHFLDM2', 'GHFLMC2', 'GHFLSDL', 'GHJGFLDM', 'GHJGFLMC', 'SFJSYD', 'ZLDWDM_1', 'ZLDWMC', 'TBMJ', 'KCMJ']


def update_attribute_value(file_type, in_path, layer_name, right_header, rel_tables, MC_tables, DLBM_values, report_file_name):
    # if file_type == DataType.shapefile:
    #     update_attribute_value_by_shapefile(in_path, layer_name, right_header, rel_tables)
    # elif file_type == DataType.fileGDB:
    #     update_attribute_value_by_fileGDB(in_path, layer_name, right_header, rel_tables, DLBM_values)

    wb = Workbook()
    output_stat_report1(wb, in_path, layer_name, MC_tables)
    wb.save(report_file_name)
    print("over")


def update_attribute_value_by_shapefile(in_path, layer_name, right_header, rel_tables):
    # layer = dataSource.GetLayer(0)
    layer = None
    dataSource = None

    try:
        start = time.time()

        wks = workspaceFactory().get_factory(DataType.shapefile)
        dataSource = wks.openFromFile(in_path, 1)
        layer = dataSource.GetLayer(0)

        layerDefn = layer.GetLayerDefn()

        field_names = []
        for i in range(layerDefn.GetFieldCount()):
            fieldName = layerDefn.GetFieldDefn(i).GetName()
            field_names.append(fieldName)

        log.info("第1步: 根据规则表的DLBM右侧表头增加矢量图层{}中的相应字段...".format(layer_name))
        for header_value in right_header:
            if header_value not in field_names:
                new_field = ogr.FieldDefn(header_value, ogr.OFTString)
                new_field.SetWidth(200)
                layer.CreateField(new_field, True)
                del new_field

        iprop = 1
        total_count = layer.GetFeatureCount()

        feature = layer.GetNextFeature()

        log.info("第2步: 根据规则表更新{}图层对应数据...".format(layer_name))
        icount = 0

        while feature:
            DLBM_value = feature.GetField("DLBM")
            bchecked = False

            for i in range(len(rel_tables)):
                rel = rel_tables[i]
                field_name = right_header[i]
                if DLBM_value not in rel:
                    if not bchecked:
                        log.warning("第{}个要素的地类编码{}在规则表中不存在！".format(icount, DLBM_value))
                        feature.SetField(field_name, None)
                        bchecked = True
                    else:
                        feature.SetField(field_name, None)
                else:
                    if feature.GetFieldType(field_name) == ogr.OFTString:
                        feature.SetField(field_name, str(rel[DLBM_value]))
                    elif feature.GetFieldType(field_name) == ogr.OFTInteger:
                        feature.SetField(field_name, int(rel[DLBM_value]))
                    elif feature.GetFieldType(field_name) == ogr.OFTReal:
                        feature.SetField(field_name, float(rel[DLBM_value]))
                    else:
                        log.error("第{}个要素的字段{}是无法识别的数据类型. 字段类型只允许是整型、字符型或者浮点型，请调整原始数据!".format(icount, field_name))
                        feature.SetField(field_name, None)

            # 如果是CZCSXM是201或者202则重新赋值
            CZCSXM_index = feature.GetFieldIndex("CZCSXM")
            if CZCSXM_index > 0:
                CZCSXM_value = feature.GetField(CZCSXM_index)
                if str(CZCSXM_value).strip() == '201' or str(CZCSXM_value).strip() == '202':
                    feature.SetField("XZFLSDL", "农用地/未利用地")
                    feature.SetField("GHJGFLDM", "07")
                    feature.SetField("GHJGFLMC", "城乡建设用地")
                    feature.SetField("SFJSYD", "是")

            layer.SetFeature(feature)
            feature = layer.GetNextFeature()

            icount += 1
            if int(icount * 100 / total_count) == iprop * 20:
                log.info("{:.0%}已处理完成...".format(icount / total_count))
                iprop += 1

        end = time.time()
        log.info("操作完成, 总共耗时:{}秒.".format("{:.2f}".format(end-start)))
        return True
    except:
        log.error("无法更新数据！错误原因:\n{}".format(traceback.format_exc()))
        return False
    finally:
        del dataSource
        del layer
        del feature
        del wks


# 采用执行SQL方式更新数据
def update_attribute_value_by_fileGDB(in_path, layer_name, right_header, rel_tables, DLBM_values):
    layer = None
    dataSource = None
    feature = None
    wks = None

    try:
        start = time.time()

        wks = workspaceFactory().get_factory(DataType.fileGDB)
        dataSource = wks.openFromFile(in_path, 1)
        layer = dataSource.GetLayerByName(layer_name)

        layerDefn = layer.GetLayerDefn()

        field_names = []
        for i in range(layerDefn.GetFieldCount()):
            fieldName = layerDefn.GetFieldDefn(i).GetName()
            field_names.append(fieldName)

        log.info("第1步: 根据规则表的DLBM右侧表头增加矢量图层{}中的相应字段...".format(layer_name))
        for header_value in right_header:
            if header_value not in field_names:
                new_field = ogr.FieldDefn(header_value, ogr.OFTString)
                new_field.SetWidth(200)
                layer.CreateField(new_field, True)
                del new_field

        log.info("第2步: 对矢量图层{}的DLBM字段创建索引...".format(layer_name))
        exec_str = r"CREATE INDEX DLBM_index ON {} (DLBM)".format(layer_name)
        dataSource.ExecuteSQL(exec_str)

        log.info("第3步: 计算矢量图层{}的DLBM字段的唯一值...".format(layer_name))
        exec_str = r"SELECT DISTINCT DLBM FROM {}".format(layer_name)
        res = dataSource.ExecuteSQL(exec_str, dialect="SQLite")

        DLBM_keys = []
        feature = res.GetNextFeature()
        while feature:
            DLBM_key = feature.GetField(0)
            DLBM_keys.append(DLBM_key)
            feature = res.GetNextFeature()

        feature = layer.GetFeature(1)

        log.info("第4步: 根据规则表计算矢量图层相应字段的值...")
        for DLBM_key in DLBM_keys:
            log.info("更新矢量图层{}字段DLBM中所有等于{}的值".format(layer_name, DLBM_key))

            if DLBM_key not in DLBM_values:
                log.warning("字段DLBM中包含规则表中不存在的编码{}".format(DLBM_key))

            for i in range(len(rel_tables)):
                rel = rel_tables[i]
                field_name = right_header[i]

                if DLBM_key in rel:
                    if feature.GetFieldType(field_name) == ogr.OFTString:
                        exec_str = r"UPDATE {} SET {} = '{}' WHERE DLBM = '{}'".format(layer_name, field_name, rel[DLBM_key], DLBM_key)
                    elif feature.GetFieldType(field_name) == ogr.OFTInteger or feature.GetFieldType(field_name) == ogr.OFTReal:
                        exec_str = r"UPDATE {} SET {} = {} WHERE DLBM = '{}'".format(layer_name, field_name, rel[DLBM_key], DLBM_key)
                    else:
                        log.error("图层{}的字段{}是无法识别的数据类型. 字段类型只允许是整型、字符型或者浮点型，请调整原始数据!".format(layer_name, field_name))
                        exec_str = r"UPDATE {} SET {} = NULL WHERE DLBM = '{}'".format(layer_name, field_name, DLBM_key)
                else:
                    exec_str = r"UPDATE {} SET {} = NULL WHERE DLBM = '{}'".format(layer_name, field_name, DLBM_key)

                dataSource.ExecuteSQL(exec_str)

        log.info("第5步: 特别更新字段CZCSXM等于201或202的对应字段值...".format(layer_name, DLBM_key))
        CZCSXM_index = feature.GetFieldIndex("CZCSXM")
        if CZCSXM_index > 0:
            XZFLSDL_index = feature.GetFieldIndex("XZFLSDL")
            GHJGFLDM_index = feature.GetFieldIndex("GHJGFLDM")
            GHJGFLMC_index = feature.GetFieldIndex("GHJGFLMC")
            SFJSYD_index = feature.GetFieldIndex("SFJSYD")

            if XZFLSDL_index > -1:
                exec_str = r"UPDATE {} SET XZFLSDL = '{}' WHERE CZCSXM = '201' or CZCSXM = '202'".format(layer_name, '农用地/未利用地')
                dataSource.ExecuteSQL(exec_str)
            if GHJGFLDM_index > -1:
                exec_str = r"UPDATE {} SET GHJGFLDM = '{}' WHERE CZCSXM = '201' or CZCSXM = '202'".format(layer_name, '07')
                dataSource.ExecuteSQL(exec_str)
            if GHJGFLMC_index > -1:
                exec_str = r"UPDATE {} SET GHJGFLMC = '{}' WHERE CZCSXM = '201' or CZCSXM = '202'".format(layer_name, '城乡建设用地')
                dataSource.ExecuteSQL(exec_str)
            if SFJSYD_index > -1:
                exec_str = r"UPDATE {} SET SFJSYD = '{}' WHERE CZCSXM = '201' or CZCSXM = '202'".format(layer_name, '是')
                dataSource.ExecuteSQL(exec_str)

        end = time.time()
        log.info("操作完成, 总共耗时:{}秒".format("{:.2f}".format(end-start)))
        return True
    except:
        log.error("无法更新数据！错误原因:\n{}".format(traceback.format_exc()))
        return False
    finally:
        del dataSource
        del layer
        del feature
        del wks

# 报表1 各区现状面积汇总表
def output_stat_report1(wb, in_path, layer_name, MC_tables):
    layer_name = 'DLTB_2'
    wks = workspaceFactory().get_factory(DataType.fileGDB)
    dataSource = wks.openFromFile(r'D:\Codes\oneTools\data\2020年国土变更调查年末库.gdb', 1)
    layer = dataSource.GetLayerByName(layer_name)

    all_field_names = check_field(dataSource, layer)

    if all_field_names is None:
        return

    exec_str = r"CREATE INDEX DLBM_index ON {} (DLBM)".format(layer_name)
    dataSource.ExecuteSQL(exec_str)
    exec_str = r"CREATE INDEX ZLDWDM_index ON {} (ZLDWDM)".format(layer_name)
    dataSource.ExecuteSQL(exec_str)
    exec_str = r"CREATE INDEX XZFLSDL_index ON {} (XZFLSDL)".format(layer_name)
    dataSource.ExecuteSQL(exec_str)


    header_font = Font(bold=True, size=11)
    header_font2 = Font(bold=True, size=9)
    header_font3 = Font(bold=False, size=11)
    cell_font = Font(bold=False, size=9)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    alignment_center = Alignment(horizontal="center", vertical="center", wrapText=True)
    alignment_right = Alignment(horizontal="right", vertical="center", wrapText=True)

    # 大表头样式
    header_style = NamedStyle(name="header_style")
    # header_style.border = border_thin
    header_style.alignment = alignment_center
    header_style.font = header_font

    # 小表头样式
    header_style2 = NamedStyle(name="header_style2")
    header_style2.border = border_thin
    header_style2.alignment = alignment_center
    header_style2.font = header_font2

    # 数据单元格居中样式
    cell_center_style = NamedStyle(name="cell_center_style")
    cell_center_style.border = border_thin
    cell_center_style.alignment = alignment_center
    cell_center_style.font = cell_font
    cell_center_style.number_format = "0.00"

    # 数据单元格居右样式
    cell_right_style = NamedStyle(name="cell_right_style")
    cell_right_style.border = border_thin
    cell_right_style.alignment = alignment_right
    cell_right_style.font = cell_font
    cell_right_style.number_format = "0.00"

    # 数据单元格居右加粗样式
    cell_right_thick_style = NamedStyle(name="cell_right_thick_style")
    cell_right_thick_style.border = border_thin
    cell_right_thick_style.alignment = alignment_right
    cell_right_thick_style.font = header_font2
    cell_right_thick_style.number_format = "0.00"

    ws = wb.create_sheet('各区现状分类面积汇总表')

    region_names = ['深圳市', '罗湖区', '福田区', '南山区', '宝安区', '龙岗区', '盐田区', '龙华区', '坪山区', '光明区', '大鹏新区']
    region_codes = ['4403', '440303', '440304', '440305', '440306', '440307', '440308', '440309', '440310', '440311', '440312']
    col_count = len(region_names) + 2

    # 注意： 要先设置样式再合并，否则边框会出问题，这是openpyxl的Bug， 相关讨论见https://foss.heptapod.net/openpyxl/openpyxl/-/issues/365
    ws.cell(1, 1).value = "表2 各区现状分类面积汇总表"
    ws.cell(1, 1).style = header_style
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)

    ws.cell(2, col_count).value = "单位：公顷"
    ws.cell(2, col_count).font = header_font3

    ws.cell(3, 1).value = "行政区域"
    ws.cell(3, 1).style = header_style2
    ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
    #
    ws.cell(3, 2).value = "名称"
    ws.cell(3, 2).style = header_style2
    ws.cell(4, 2).value = "代码"
    ws.cell(4, 2).style = header_style2

    ws.cell(5, 1).value = "国土调查总面积"
    ws.cell(5, 1).style = header_style2
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=2)

    ws.cell(6, 1).value = "三大类"
    ws.cell(6, 1).style = header_style2
    ws.merge_cells(start_row=6, start_column=1, end_row=8, end_column=1)

    ws.cell(6, 2).value = "农用地"
    ws.cell(6, 2).style = cell_center_style
    ws.cell(7, 2).value = "建设用地"
    ws.cell(7, 2).style = cell_center_style
    ws.cell(8, 2).value = "未利用地"
    ws.cell(8, 2).style = cell_center_style

    for i in range(3, col_count + 1):
        ws.cell(3, i).value = region_names[i - 3]
        ws.cell(3, i).style = header_style2

        ws.cell(4, i).value = region_codes[i - 3]
        ws.cell(4, i).style = cell_center_style

        ws.cell(5, i).style = header_style2

        # exec_str = r"SELECT SUM(TBMJ) FROM {} WHERE XZFLSDL='{}' AND SUBSTR(ZLDWDM, 1, 4)='{}'".format(layer_name, ws.cell(6, 2).value, region_codes[i - 3])
        # exec_layer = dataSource.ExecuteSQL(exec_str, dialect="SQLite")
        # exec_res = exec_layer.GetNextFeature().GetField(0)
        # ws.cell(6, i).value = exec_res

    # 统计三大类面积
    for i in range(0, 3):
        exec_str = r"SELECT SUBSTR(ZLDWDM_1, 1, 6), SUM(TBMJ) FROM {} WHERE XZFLSDL='{}' GROUP BY SUBSTR(ZLDWDM_1, 1, 6)".format(layer_name, ws.cell(6 + i, 2).value)
        exec_layer = dataSource.ExecuteSQL(exec_str, dialect="SQLite")

        if exec_layer is None:
            log.warning("{}执行结果为空!".format(exec_str))
        else:
            exec_res = exec_layer.GetNextFeature()

            while exec_res:
                ZDDWDM = exec_res.GetField(0)
                ZDDWDM_MJ = '%.2f' % (exec_res.GetField(1) / 10000)

                pos = region_codes.index(ZDDWDM)
                if pos > -1:
                    ws.cell(6 + i, pos + 3).value = float(ZDDWDM_MJ)
                    ws.cell(6 + i, pos + 3).style = cell_right_style
                else:
                    log.warning("没有相应的区域代码{}!".format(ZDDWDM))

                exec_res = exec_layer.GetNextFeature()

        # # 深圳市的总计
        # sum = 0
        # for iRange in range(4, col_count + 1):
        #     sum = sum + ws.cell(6 + i, iRange)
        # ws.cell(6 + i, 3).value = float(sum)
        # ws.cell(6 + i, 3).style = cell_right_style

    i = 0
    start_row = 9
    for mc, DLMCs in MC_tables.items():
        ws.cell(start_row, 1).value = "{}\n({})".format(mc, str(i).zfill(2))
        ws.cell(start_row, 1).style = header_style2

        for j in range(len(DLMCs)):
            ws.cell(start_row, 2).value = "小计\n({})".format(str(i).zfill(2))
            ws.cell(start_row, 2).style = header_style2
            ws.cell(start_row + j + 1, 2).value = "{}\n({})".format(DLMCs[j]["DLMC"], DLMCs[j]["DLBM"])
            ws.cell(start_row + j + 1, 2).style = cell_center_style

            for iRegion in range(3, col_count + 1):
                ws.cell(start_row + j + 1, iRegion).value = float('%.2f' % 0.00)
                ws.cell(start_row + j + 1, iRegion).style = cell_right_style

            DLBM_key = DLMCs[j]["DLBM"]
            exec_str = r"SELECT SUBSTR(ZLDWDM_1, 1, 6), SUM(TBMJ) FROM {} WHERE DLBM='{}' GROUP BY SUBSTR(ZLDWDM_1, 1, 6)".format(layer_name, DLBM_key)
            exec_layer = dataSource.ExecuteSQL(exec_str, dialect="SQLite")

            if exec_layer is None:
                log.warning("{}执行结果为空!".format(exec_str))
            else:
                exec_res = exec_layer.GetNextFeature()

                while exec_res:
                    ZDDWDM = exec_res.GetField(0)
                    ZDDWDM_MJ = '%.2f' % (exec_res.GetField(1) / 10000)

                    pos = region_codes.index(ZDDWDM)
                    if pos > -1:
                        ws.cell(start_row + j + 1, pos + 3).value = float(ZDDWDM_MJ)
                    else:
                        log.warning("没有相应的区域代码{}!".format(ZDDWDM))
                    ws.cell(start_row + j + 1, pos + 3).style = cell_right_style

                    exec_res = exec_layer.GetNextFeature()

        # 二级分类小计
        for iRegion in range(3, col_count + 1):
            sum = 0
            for j in range(len(DLMCs)):
                sum = sum + float(ws.cell(start_row + j + 1, iRegion).value)
            ws.cell(start_row, iRegion).value = float('%.2f' % sum)
            ws.cell(start_row, iRegion).style = cell_right_thick_style

        if len(DLMCs) > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row+len(DLMCs), end_column=1)
            start_row = start_row + len(DLMCs) + 1  # 增加一行"小计"
        else:
            start_row = start_row + len(DLMCs)
        i += 1

    # 深圳市总计
    for i in range(6, start_row):
        sum = 0
        for iRange in range(4, col_count + 1):
            sum = sum + float(ws.cell(i, iRange).value)
        ws.cell(i, 3).value = float(sum)
        ws.cell(i, 3).style = cell_right_thick_style


def readSpatialData(file_type, in_path, layer_name):
    if file_type == DataType.shapefile:
        wks = workspaceFactory().get_factory(DataType.shapefile)
        datasource = wks.openFromFile(in_path, 1)
        layer = datasource.GetLayer(0)

    elif file_type == DataType.fileGDB:
        wks = workspaceFactory().get_factory(DataType.fileGDB)
        datasource = wks.openFromFile(in_path, 1)
        layer = datasource.GetLayerByName(layer_name)

    return layer


def check_field(dataSource, layer):
    all_field_names = []
    layer_name = layer.GetName()

    berror = False
    layerDefn = layer.GetLayerDefn()
    for i in range(layerDefn.GetFieldCount()):
        fieldName = layerDefn.GetFieldDefn(i).GetName()
        if fieldName.upper() == 'ZLDWDM':
            all_field_names.append('ZLDWDM_1')
        else:
            all_field_names.append(fieldName.upper())

    if layerDefn.GetFieldIndex('ZLDWDM_1') < 0:
        new_field = ogr.FieldDefn('ZLDWDM_1', ogr.OFTString)
        new_field.SetWidth(200)
        layer.CreateField(new_field, True)
        del new_field

    for need_field in report_need_fields:
        if need_field not in all_field_names:
            log.warning('缺失输出报表得必要字段"{}"，无法执行输出报表操作，请补全！'.format(need_field))
            berror = True

    exec_str = r"UPDATE {} SET ZLDWDM_1=ZLDWDM WHERE 1=1".format(layer_name)
    dataSource.ExecuteSQL(exec_str)
    exec_str = r"UPDATE {} SET ZLDWDM_1='4403120000000000000' WHERE ZLDWDM LIKE '440307%' AND " \
               r"ZLDWMC <> '宝龙街道' AND ZLDWMC <> '布吉街道' AND ZLDWMC <> '龙城街道' AND " \
               r"ZLDWMC <> '龙岗街道' AND ZLDWMC <> '平湖街道' AND ZLDWMC <> '坪地街道' AND " \
               r"ZLDWMC <> '园山街道' AND ZLDWMC <> '南湾街道' AND ZLDWMC <> '坂田街道' AND " \
               r"ZLDWMC <> '吉华街道' AND ZLDWMC <> '横岗街道'".format(layer_name)
    dataSource.ExecuteSQL(exec_str)

    if berror:
        return None
    else:
        return all_field_names
