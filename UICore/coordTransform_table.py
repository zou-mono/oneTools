import csv
import time
import traceback

import click
import os

from openpyxl import load_workbook
from osgeo import ogr, osr, gdal

from UICore.DataFactory import workspaceFactory, get_row_from_excel
from UICore.Gv import DataType, DataType_dict, srs_dict
from UICore.common import launderName, overwrite_cpg_file, is_already_opened_in_write_mode, \
    helmert_para_dict, get_suffix, is_number, text_line_count
from UICore.coordTransform_dwg import transform_dwg
from UICore.coordTransform_web import gcj02_to_wgs84_acc, wgs84_to_gcj02, bd09_to_wgs84_acc, wgs84_to_bd09, \
    gcj02_to_wgs84_acc_list, wgs84_to_gcj02_list, bd09_to_wgs84_acc_list, wgs84_to_bd09_list
from UICore.log4p import Log
from UICore.Gv import SpatialReference

log = Log(__file__)

@click.command()
@click.option(
    '--inpath', '-i',
    help='Input table file. For example, d:/res/data/xxx.csv',
    type=str,
    required=True)
@click.option(
    '--inencode',
    help='Input encoding. The default is utf-8',
    type=str,
    default='utf-8',
    required=False)
@click.option(
    '--header', '-h',
    help='Header exist or not, the default is yes.',
    type=bool,
    default=True,
    required=False)
@click.option(
    '--xfield', '-x',
    help='The order of x field.',
    type=int,
    required=True)
@click.option(
    '--yfield', '-y',
    help='The order of y field.',
    type=int,
    required=True)
@click.option(
    '--insrs',
    help='Input srs. sz_Local = 2435, gcs_2000 = 4490, pcs_2000 = 4547, pcs_2000_zone = 4526, wgs84 = 4326, bd09 = -1, '
         'gcj02 = -2, gcs_xian80 = 4610, pcs_xian80 = 2383, pcs_xian80_zone = 2363. The in layer\'s srs will be used to the default',
    type=int,
    default='-99',
    required=False)
@click.option(
    '--outsrs',
    help='Output srs. sz_Local = 0, gcs_2000 = 1, pcs_2000 = 2, pcs_2000_zone = 3, wgs84 = 4, bd09 = 5, '
         'gcj02 = 6, gcs_xian80 = 7, pcs_xian80 = 8, pcs_xian80_zone = 9.',
    type=int,
    required=True)
@click.option(
    '--outpath', '-o',
    help='Output table file. For example, d:/res/data/xxx.csv',
    type=str,
    required=True)
@click.option(
    '--outencode',
    help='Output encoding. For example, d:/res/data/xxx.csv',
    type=str,
    default='gbk',
    required=False)
def main(inpath, inencode, header, xfield, yfield, insrs, outsrs, outpath, outencode):
    """spatial coordinate transformation program"""
    coordTransform(inpath, inencode, header, xfield, yfield, insrs, outsrs, outpath, outencode)

def coordTransform(inpath, inencode, header, xfield, yfield, insrs, outsrs, outpath, outencode):
    gdal.SetConfigOption("OGR_CT_FORCE_TRADITIONAL_GIS_ORDER", "YES")

    if inpath[-1] == os.sep:
        inpath = inpath[:-1]
    if outpath[-1] == os.sep:
        outpath = outpath[:-1]

    in_format = get_suffix(inpath)
    out_format = get_suffix(outpath)

    try:
        tfer = Transformer(in_format, out_format, inpath, inencode, header, xfield, yfield, outpath, outencode)
        tfer.transform(insrs, outsrs)
        return True, ''
    except:
        return False, traceback.format_exc()


class Transformer(object):
    def __init__(self, informat, outformat, inpath, inencode, header, x, y, outpath, outencode):
        self.out_format = outformat
        self.in_format = informat
        self.in_path = inpath
        self.in_encode = inencode
        self.header = header
        self.out_encode = outencode
        self.out_path = outpath
        self.x = x
        self.y = y

        self.rowcount = 0  # 记录遍历到的行号
        self.iprop = 1  # 用于百分比输出
        self.points = []  # 记录缓存的点集
        self.rows = []  # 记录缓存的行
        self.fail_rows = []  # 记录下无法转换的行号
        self.total_count = 0  # 记录总行数

    def transform(self, srcSRS, dstSRS):
        self.srcSRS = srcSRS
        self.dstSRS = dstSRS

        start = time.time()

        res = None
        if self.in_format == DataType.csv:
            export_func = self.export_csv_to_file
        elif self.in_format == DataType.xlsx:
            export_func = self.export_xlsx_to_file
        elif self.in_format == DataType.dbf:
            export_func = self.export_dbf_to_file
        else:
            return False

        log.info("启动从{}到{}的转换...".format(self.in_path, self.out_path))

        if srcSRS == SpatialReference.sz_Local and dstSRS == SpatialReference.pcs_2000:
            res = export_func(self.sz_local_to_pcs_2000)
        elif srcSRS == SpatialReference.pcs_2000 and dstSRS == SpatialReference.sz_Local:
            res = export_func(self.pcs_2000_to_sz_local)
        elif srcSRS == SpatialReference.sz_Local and dstSRS == SpatialReference.gcs_2000:
            res = export_func(self.sz_local_to_gcs_2000)
        elif srcSRS == SpatialReference.gcs_2000 and dstSRS == SpatialReference.sz_Local:
            res = export_func(self.gcs_2000_to_sz_local)
        elif srcSRS == SpatialReference.sz_Local and dstSRS == SpatialReference.wgs84:
            res = export_func(self.sz_local_to_wgs84)
        elif srcSRS == SpatialReference.wgs84 and dstSRS == SpatialReference.sz_Local:
            res = export_func(self.wgs84_to_sz_local)
        elif srcSRS == SpatialReference.sz_Local and dstSRS == SpatialReference.pcs_2000_zone:
            res = export_func(self.sz_local_to_pcs_2000_zone)
        elif srcSRS == SpatialReference.pcs_2000_zone and dstSRS == SpatialReference.sz_Local:
            res = export_func(self.pcs_2000_zone_to_sz_local)
        elif srcSRS == SpatialReference.wgs84 and dstSRS == SpatialReference.pcs_2000:
            res = export_func(self.wgs84_to_pcs_2000)
        elif srcSRS == SpatialReference.pcs_2000_zone and dstSRS == SpatialReference.wgs84:
            res = export_func(self.pcs_2000_zone_to_wgs84)
        elif srcSRS == SpatialReference.pcs_2000 and dstSRS == SpatialReference.wgs84:
            res = export_func(self.pcs_2000_to_wgs84)
        elif srcSRS == SpatialReference.wgs84 and dstSRS == SpatialReference.pcs_2000_zone:
            res = export_func(self.wgs84_to_pcs_2000_zone)
        elif srcSRS == SpatialReference.gcj02 and dstSRS == SpatialReference.wgs84:
            res = export_func(self.gcj02_to_wgs84)
        elif srcSRS == SpatialReference.wgs84 and dstSRS == SpatialReference.gcj02:
            res = export_func(self.wgs84_gcj02)
        elif srcSRS == SpatialReference.bd09 and dstSRS == SpatialReference.wgs84:
            res = export_func(self.bd09_to_wgs84)
        elif srcSRS == SpatialReference.wgs84 and dstSRS == SpatialReference.bd09:
            res = export_func(self.wgs84_to_bd09)
        elif srcSRS == SpatialReference.gcj02 and dstSRS == SpatialReference.sz_Local:
            res = export_func(self.gcj02_to_sz_local)
        elif srcSRS == SpatialReference.bd09 and dstSRS == SpatialReference.sz_Local:
            res = export_func(self.bd09_to_sz_local)
        elif srcSRS == SpatialReference.gcj02 and dstSRS == SpatialReference.pcs_2000:
            res = export_func(self.gcj02_to_pcs_2000)
        elif srcSRS == SpatialReference.bd09 and dstSRS == SpatialReference.pcs_2000:
            res = export_func(self.bd09_to_pcs_2000)
        else:
            log.error("不支持从{}到{}的转换!".format(srs_dict[srcSRS], srs_dict[dstSRS]))
            return False

        end = time.time()

        if res is None:
            log.info("坐标转换完成! 共耗时{}秒. 输出路径:{}"
                     .format("{:.2f}".format(end-start), self.out_path))
        else:
            log.error("坐标转换失败!\n{}".format(res))

    def export_csv_to_file(self, transform_method):
        self.rowcount = 0  # 记录遍历到的行号
        self.iprop = 1  # 用于百分比输出
        self.points = []  # 记录缓存的点集
        self.rows = []  # 记录缓存的行
        self.fail_rows = []  # 记录下无法转换的行号
        self.total_count = 0  # 记录总行数

        try:
            self.total_count = text_line_count(self.in_path, self.in_encode)

            with open(self.in_path, "r", encoding=self.in_encode) as f:
                reader = csv.reader(f)
                if self.header:
                    header = next(reader)
                    self.total_count = self.total_count - 1

                with open(self.out_path, 'w+', encoding=self.out_encode, newline='') as o:
                    writer = csv.writer(o)
                    if self.header:
                        header.extend(["{}_x".format(srs_dict[self.dstSRS]),
                                       "{}_y".format(srs_dict[self.dstSRS])])
                        writer.writerow(header)

                    for row in reader:
                        self.rowcount += 1
                        self.write_row_to_csv(row, writer=writer, transform_method=transform_method)

                return None
        except:
            return traceback.format_exc()

    def export_xlsx_to_file(self, transform_method):
        self.rowcount = 0  # 记录遍历到的行号
        self.iprop = 1  # 用于百分比输出
        self.points = []  # 记录缓存的点集
        self.rows = []  # 记录缓存的行
        self.fail_rows = []  # 记录下无法转换的行号
        self.total_count = 0  # 记录总行数

        wb = None
        try:
            wb = load_workbook(self.in_path, read_only=True)
            ws = wb.active
            rows = ws.rows

            # columns = ws.max_column
            self.total_count = ws.max_row

            with open(self.out_path, 'w+', encoding=self.out_encode, newline='') as o:
                writer = csv.writer(o)

                if self.header:
                    header = []
                    excel_row = next(rows)
                    for cell in excel_row:
                        header.append(str(cell.value))
                    # header = get_row_from_excel(ws, 1, columns)
                    header.extend(["{}_x".format(srs_dict[self.dstSRS]),
                                   "{}_y".format(srs_dict[self.dstSRS])])
                    writer.writerow(header)
                    self.total_count = self.total_count - 1

                for excel_row in rows:
                    self.rowcount += 1

                    row = []
                    # row = get_row_from_excel(ws, self.rowcount, columns)
                    for cell in excel_row:
                        row.append(str(cell.value))
                    self.write_row_to_csv(row, writer=writer, transform_method=transform_method)

            return None
        except:
            return traceback.format_exc()
        finally:
            if wb is not None:
                wb.close()

    def export_dbf_to_file(self, transform_method):
        wks = workspaceFactory().get_factory(DataType.dbf)
        datasource = wks.openFromFile(self.in_path)
        if datasource is not None:
            layer = datasource.GetLayer()
            for row in layer:
                print(row)

    def write_row_to_csv(self, row, writer, transform_method):
        if not is_number(row[self.x]) or not is_number(row[self.y]):
            log.warning("第{}行的坐标包含非数字字段，无法转换!".format(self.rowcount))
            # writer.writerow([])
            self.fail_rows.append(self.rowcount)
            return False

        self.points.append([float(row[0]), float(row[1])])
        self.rows.append(row)

        if self.rowcount % 5000 == 0 or self.rowcount == self.total_count:
            points = transform_method(self.points)
            for fail_row in self.fail_rows:
                points.insert(fail_row - 1, [])
                self.rows.insert(fail_row - 1, [])
            for i in range(len(points)):
                temp_row = self.rows[i]
                if len(points[i]) > 0:
                    temp_row.extend([points[i][0], points[i][1]])
                else:
                    temp_row = []
                writer.writerow(temp_row)
            # writer.writerows(points)
            self.rows = []
            self.points = []
            self.fail_rows = []

        if int(self.rowcount * 100 / self.total_count) == self.iprop * 20:
            log.debug("{:.0%}".format(self.rowcount / self.total_count))
            self.iprop += 1

        return True


    def proj_transform(self, srcSRS, dstSRS, points):
        sourceSRS = osr.SpatialReference()
        sourceSRS.ImportFromEPSG(srcSRS)
        destinationSRS = osr.SpatialReference()
        destinationSRS.ImportFromEPSG(dstSRS)
        tr = osr.CreateCoordinateTransformation(sourceSRS, destinationSRS)
        points = osr.CoordinateTransformation.TransformPoints(tr, points)
        return points

    # 关键转换，需要参数
    def sz_local_to_pcs_2000(self, points):
        helmert_para = helmert_para_dict(2435, 4547, "EAST")
        opt = osr.CoordinateTransformationOptions()
        opt.SetOperation(helmert_para)
        tr = osr.CreateCoordinateTransformation(None, None, opt)
        points = osr.CoordinateTransformation.TransformPoints(tr, points)
        return points

    # 关键转换，需要参数
    def pcs_2000_to_sz_local(self, points):
        helmert_para = helmert_para_dict(4547, 2435, "EAST")
        opt = osr.CoordinateTransformationOptions()
        opt.SetOperation(helmert_para)
        tr = osr.CreateCoordinateTransformation(None, None, opt)
        points = osr.CoordinateTransformation.TransformPoints(tr, points)
        return points

    def sz_local_to_gcs_2000(self, points):
        points = self.sz_local_to_pcs_2000(points)
        points = self.proj_transform(4547, 4490, points)
        # sourceSRS = osr.SpatialReference()
        # sourceSRS.ImportFromEPSG(4547)
        # destinationSRS = osr.SpatialReference()
        # destinationSRS.ImportFromEPSG(4490)
        # tr = osr.CreateCoordinateTransformation(sourceSRS, destinationSRS)
        # # pt = osr.CoordinateTransformation.TransformPoints(tr, [[434431.2239, 431392.0434], [431392.0434, 2607747.97677669]])
        # # [434431.2239, 431392.0434, 431408.4703, 412476.2589, 412535.6403],
        # #                              [2612265.488, 2607747.97677669, 2607774.43119491, 2592367.01357008, 2592864.81664873]
        # # point.Transform(tr)
        # points = osr.CoordinateTransformation.TransformPoints(tr, points)
        return points

    def gcs_2000_to_sz_local(self, points):
        points = self.proj_transform(4490, 4547, points)
        points = self.pcs_2000_to_sz_local(points)
        return points

    def sz_local_to_wgs84(self, points):
        points = self.sz_local_to_pcs_2000(points)
        points = self.proj_transform(4547, 4326, points)
        return points

    def wgs84_to_sz_local(self, points):
        points = self.proj_transform(4326, 4547, points)
        points = self.pcs_2000_to_sz_local(points)
        return points

    def sz_local_to_pcs_2000_zone(self, points):
        points = self.sz_local_to_pcs_2000(points)
        points = self.proj_transform(4547, 4526, points)
        return points

    def pcs_2000_zone_to_sz_local(self, points):
        points = self.proj_transform(4526, 4547, points)
        points = self.pcs_2000_to_sz_local(points)
        return points

    def wgs84_to_pcs_2000(self, points):
        points = self.proj_transform(4326, 4547, points)
        return points

    def pcs_2000_to_wgs84(self, points):
        points = self.proj_transform(4547, 4326, points)
        return points

    def wgs84_to_pcs_2000_zone(self, points):
        points = self.proj_transform(4326, 4526, points)
        return points

    def pcs_2000_zone_to_wgs84(self, points):
        points = self.proj_transform(4526, 4326, points)
        return points

    def gcj02_to_wgs84(self, points):
        points = list(map(gcj02_to_wgs84_acc_list, points))
        return points

    def wgs84_gcj02(self, points):
        points = list(map(wgs84_to_gcj02_list, points))
        return points

    def bd09_to_wgs84(self, points):
        points = list(map(bd09_to_wgs84_acc_list, points))
        return points

    def wgs84_to_bd09(self, points):
        points = list(map(wgs84_to_bd09_list, points))
        return points

    def gcj02_to_sz_local(self, points):
        points = self.gcj02_to_wgs84(points)
        points = self.wgs84_to_sz_local(points)
        return points

    def bd09_to_sz_local(self, points):
        points = self.bd09_to_wgs84(points)
        points = self.wgs84_to_sz_local(points)
        return points

    def gcj02_to_pcs_2000(self, points):
        points = self.gcj02_to_wgs84(points)
        points = self.wgs84_to_pcs_2000(points)
        return points

    def bd09_to_pcs_2000(self, points):
        points = self.bd09_to_wgs84(points)
        points = self.wgs84_to_pcs_2000(points)
        return points


def launderLayerName(path):
    if is_already_opened_in_write_mode(path):
        return launderName(path)
    else:
        return path


# 获取axis order，先北后东还是先东后北
def get_axis_order(srs):
    srs_wkt = osr.SpatialReference(srs.ExportToWkt())
    order0 = srs_wkt.GetAttrValue("AXIS", 1)
    return order0


if __name__ == '__main__':
    ogr.UseExceptions()
    gdal.SetConfigOption("OGR_CT_FORCE_TRADITIONAL_GIS_ORDER", "YES")
    main()
