import base64
import csv
import json
import math
import os
import re
import datetime
import time

import chardet
from openpyxl import load_workbook
from openpyxl.styles import Border

from UICore.Gv import srs_dict, SpatialReference, DataType
from UICore.log4p import Log
import urllib.request, urllib.parse

log = Log()


def is_already_opened_in_write_mode(filename):
    if os.path.exists(filename):
        try:
            # 通过尝试修改名字是否报错来判断是否被使用
            os.rename(filename, filename)
        except IOError:
            return True
    return False


def get_json(url):
    try_num = 5
    # 定义请求头
    # reqheaders = {'Content-Type': 'application/x-www-form-urlencoded',
    #               # 'Host': 'suplicmap.pnr.sz',
    #               'Pragma': 'no-cache'}
    # 请求不同页面的数据
    trytime = 0
    while trytime < try_num:
        try:
            req = urllib.request.Request(url=url)
            r = urllib.request.urlopen(req)
            respData = r.read().decode('utf-8', 'ignore')
            # return respData
            log.debug(respData)
            res = json.loads(respData)
            if 'error' not in res.keys():
                return res
            else:
                trytime += 1
        except:
            # log.error('HTTP请求失败！重新尝试...')
            trytime += 1

        time.sleep(2)
        continue


def get_paraInfo(url):
    http = r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+'
    res = re.match(http, string=url)
    log.debug(res)
    url_json = url + "?f=pjson"
    if res is not None:
        getInfo = get_json(url_json)
        return getInfo
    else:
        return None


def defaultTileFolder(url, level):
    # url_encodeStr = str(base64.b64encode(url.encode("utf-8")), "utf-8")
    url_encodeStr = urlEncodeToFileName(url)
    path = os.path.join(os.getcwd(), "data", "tiles", url_encodeStr, str(level))
    if not os.path.exists(path):
        os.makedirs(path)
    return path


def urlEncodeToFileName(url):
    url_encodeStr = str(base64.b64encode(url.encode("utf-8")), "utf-8")
    return url_encodeStr


def defaultImageFile(url, level):
    url_encodeStr = str(base64.b64encode(url.encode("utf-8")), "utf-8")
    if not os.path.exists(os.path.join(os.getcwd(), "data")):
        os.makedirs(os.path.join(os.getcwd(), "data"))
    file = os.path.join(os.getcwd(), "data", "{}_{}".format(url_encodeStr, str(level)))
    return launderName(file + '.tif')


def launderName(name):
    dir = os.path.dirname(name)
    basename, suffix = os.path.splitext(name)
    if os.path.exists(name):
        basename = basename + "_1"
        name = os.path.join(dir, basename + suffix)

    if not (os.path.exists(name)):
        return name
    else:
        return launderName(name)


def get_col_row(x0, y0, x, y, size, resolution):
    col = math.floor(math.fabs((x0 - x) / (size * resolution)))
    row = math.floor(math.fabs((y0 - y) / (size * resolution)))

    return col, row


def check_layer_name(name):
    p1 = r'[-!&<>"\'?@=$~^`#%*()/\\:;{}\[\]|+.]'
    res = re.sub(p1, '_', name)
    p2 = r'( +)'
    return re.sub(p2, '', res)
# def get_srs_desc_by_epsg(name: str):
#     if name == "2435":
#         return srs_dict[SpatialReference.sz_Local]
#     elif name == "4490":
#         return srs_dict[SpatialReference.gcs_2000]
#     elif name == "4547":
#         return srs_dict[SpatialReference.pcs_2000]
#     elif name == "4526":
#         return srs_dict[SpatialReference.pcs_2000_zone]
#     elif name == "4326":
#         return srs_dict[SpatialReference.wgs84]
#     elif name == "4610":
#         return srs_dict[SpatialReference.gcs_xian80]
#     elif name == "2383":
#         return srs_dict[SpatialReference.pcs_xian80]
#     elif name == "2362":
#         return srs_dict[SpatialReference.pcs_xian80_zone]


def overwrite_cpg_file(outpath, outfile, encoding):
    f = None
    try:
        cpg_file = os.path.join(outpath, outfile + ".cpg")
        if not os.path.exists(outpath):
            os.makedirs(outpath)

        with open(cpg_file, "w+") as f:
            f.seek(0)
            f.truncate() #清空文件
            f.write(encoding)
    finally:
        if f is not None:
            f.close()


def helmert_para_dict(insrs, outsrs, first_order="NORTH"):
    if insrs == SpatialReference.sz_Local and outsrs == SpatialReference.pcs_2000:
        if first_order == "NORTH":
            return "+proj=helmert +convention=position_vector +x={} +y={} +s={} +theta={}".format(
                2472660.600279, 391090.578943, 0.999997415382, 3518.95267316)
        else:
            return "+proj=helmert +convention=position_vector +x={} +y={} +s={} +theta={}".format(
                391090.578943, 2472660.600279, 0.999997415382, -3518.95267316)
    elif insrs == SpatialReference.pcs_2000 and outsrs == SpatialReference.sz_Local:
        if first_order == "NORTH":
            return "+proj=helmert +convention=position_vector +x={} +y={} +s={} +theta={}".format(
                -2465635.316383, -433217.228947, 1.000002584625, -3518.95267316)
        else:
            return "+proj=helmert +convention=position_vector +x={} +y={} +s={} +theta={}".format(
                -433217.228947, -2465635.316383, 1.000002584625, 3518.95267316)
    elif insrs == SpatialReference.pcs_xian80 and outsrs == SpatialReference.sz_Local:
        if first_order == "NORTH":
            return "+proj=helmert +convention=position_vector +x={} +y={} +s={} +theta={}".format(
                -2465659.407210, -433097.707045, 1.000009894628, -3518.45262840)
        else:
            return "+proj=helmert +convention=position_vector +x={} +y={} +s={} +theta={}".format(
                -433097.707045, -2465659.407210, 1.000009894628, 3518.45262840)
    else:
        return None


def get_suffix(path):
    suffix = None
    basename = os.path.basename(path)
    filename, suffix = os.path.splitext(path)
    # if basename.find('.') > 0:
    #     suffix = basename.split('.')[1]

    if suffix is None:
        return None

    if suffix.lower() == '.shp':
        return DataType.shapefile
    elif suffix.lower() == '.geojson':
        return DataType.geojson
    elif suffix.lower() == '.gdb':
        return DataType.fileGDB
    elif suffix.lower() == '.dwg':
        return DataType.cad_dwg
    elif suffix.lower() == ".csv":
        return DataType.csv
    elif suffix.lower() == ".dbf":
        return DataType.dbf
    elif suffix.lower() == ".xlsx":
        return DataType.xlsx
    else:
        return None


def encodeCurrentTime():
    # localtime = time.localtime(time.time())
    # str_time = time.asctime(localtime)
    # encode_time = str(base64.b64encode(str_time.encode("utf-8")), "utf-8")
    encode_time = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    return encode_time


# 判断第一行是否是表头
def is_header(line):
    return not any(cell.replace(".", "").isdigit() for cell in line)


def is_number(str):
    return str.replace(".", "").isdigit()


def text_line_count(in_path, in_encode):
    with open(in_path, "r", encoding=in_encode) as f:
        total_count = sum(1 for row in f)
    return total_count


def check_encoding(file):
    with open(file, 'rb') as f:
        data = f.read(10000)  # or a chunk, f.read(1000000)
        encoding = chardet.detect(data).get("encoding")

    return encoding


def read_first_line(file, format, sheet=None, encoding=None):
    header = []
    if format == DataType.csv:
        with open(file, 'r', newline='', encoding=encoding) as f:
            reader = csv.reader(f)
            header = next(reader)  # gets the first line
    elif format == DataType.xlsx:
        wb = load_workbook(file, read_only=True)
        if sheet is None:
            ws = wb.active
        else:
            ws = wb[sheet]
        # ws = wb.get_sheet_by_name(sheet)
        columns = ws.max_column
        header = []
        for i in range(1, columns + 1):
            cell_value = ws.cell(row=1, column=i).value
            header.append(str(cell_value))
        wb.close()
    elif format == DataType.memory:
        for line in file.splitlines():
            l_arr = line.split('\t')
            header = l_arr
            break

    return header


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill