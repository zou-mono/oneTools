import csv
import traceback

from PyQt5.QtCore import QRect, Qt, QPersistentModelIndex, QItemSelectionModel, QModelIndex, QThread, QObject
from PyQt5.QtGui import QDoubleValidator, QIntValidator, QPalette
from PyQt5.QtWidgets import QApplication, QDialog, QMessageBox, QErrorMessage, QDialogButtonBox, QStyleFactory, \
    QAbstractItemView, QHeaderView, QComboBox, QAbstractButton, QFileDialog
from PyQt5 import QtWidgets, QtGui, QtCore
from openpyxl import load_workbook
from osgeo import osr, gdal

import UI.UICoordTransform
import UI.listview_dialog
import UICore.coordTransform_table
import sys
import json
import os

from UI.UICoordTransform import Ui_Dialog

from UICore.DataFactory import workspaceFactory, read_table_header
from UICore.Gv import SplitterState, Dock, DataType, srs_dict
from UICore.common import defaultImageFile, defaultTileFolder, urlEncodeToFileName, get_paraInfo, get_suffix, \
    encodeCurrentTime, is_header, is_already_opened_in_write_mode, launderName, check_encoding, read_first_line
from UICore.workerThread import coordTransformWorker
from widgets.mTable import TableModel, mTableStyle, layernameDelegate, srsDelegate, outputPathDelegate, xyfieldDelegate
from UICore.log4p import Log

Slot = QtCore.pyqtSlot

log = Log(__name__)


# class myFileSelect(QFileDialog):
#     def __init__(self):
#         super().__init__()

class Ui_Window(QtWidgets.QDialog, Ui_Dialog):
    def __init__(self, parent=None):
        super(Ui_Window, self).__init__(parent=parent)
        self.setupUi(self)

        font = QtGui.QFont()
        font.setFamily("微软雅黑")
        font.setPointSize(9)
        self.buttonBox.setStandardButtons(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self.buttonBox.button(QDialogButtonBox.Ok).setFont(font)
        self.buttonBox.button(QDialogButtonBox.Ok).setText("确定")
        self.buttonBox.button(QDialogButtonBox.Cancel).setFont(font)
        self.buttonBox.button(QDialogButtonBox.Cancel).setText("取消")

        vlayout = QtWidgets.QVBoxLayout(self)
        vlayout.addWidget(self.splitter)
        vlayout.setContentsMargins(0, 0, 10, 10)
        self.splitter.setGeometry(0, 0, self.width(), self.height())

        self.splitter.setOrientation(Qt.Horizontal)
        self.splitter.setProperty("Stretch", SplitterState.collapsed)
        self.splitter.setProperty("Dock", Dock.right)
        self.splitter.setProperty("WidgetToHide", self.txt_log)
        self.splitter.setProperty("ExpandParentForm", True)

        self.splitter.setSizes([600, self.splitter.width() - 590])
        self.resize(self.splitter.width(), self.splitter.height())

        self.paras = {}  # 存储参数信息
        self.selIndex = QModelIndex()
        self.table_init()

        log.setLogViewer(parent=self, logViewer=self.txt_log)
        self.txt_log.setReadOnly(True)

        self.btn_addressFile.clicked.connect(self.open_addressFile)
        self.btn_addRow.clicked.connect(self.btn_addRow_clicked)
        self.btn_removeRow.clicked.connect(self.btn_removeBtn_clicked)
        self.btn_saveMetaFile.clicked.connect(self.btn_saveMetaFile_clicked)
        self.buttonBox.clicked.connect(self.buttonBox_clicked)
        self.splitter.splitterMoved.connect(self.splitterMoved)
        self.splitter.handle(1).handleClicked.connect(self.handleClicked)

        self.rbtn_file.clicked.connect(self.rbtn_toggled)
        self.rbtn_filedb.clicked.connect(self.rbtn_toggled)
        self.rbtn_table.clicked.connect(self.rbtn_toggled)

        self.thread = QThread(self)
        self.coordTransformThread = coordTransformWorker()
        self.coordTransformThread.moveToThread(self.thread)
        self.coordTransformThread.transform.connect(self.coordTransformThread.coordTransform)
        self.coordTransformThread.transform_tbl.connect(self.coordTransformThread.tableTransform)
        self.coordTransformThread.finished.connect(self.threadStop)

        self.bInit = True

    def showEvent(self, a0: QtGui.QShowEvent) -> None:
        log.setLogViewer(parent=self, logViewer=self.txt_log)
        if self.bInit:
            self.rbtn_file.click()
            self.bInit = False

        # self.table_layout()

    def resizeEvent(self, a0: QtGui.QResizeEvent) -> None:
        self.table_layout()

    def threadStop(self):
        def threadStop(self):
            if self.thread.isRunning():
                self.thread.terminate()
                self.thread.wait()
            else:
                self.thread.quit()

    def splitterMoved(self):
        self.table_layout()

    def handleClicked(self):
        self.table_layout()

    def table_layout(self):
        if self.rbtn_table.isChecked():
            self.tbl_address.setColumnWidth(0, self.tbl_address.width() * 0.2)
            self.tbl_address.setColumnWidth(1, self.tbl_address.width() * 0.1)
            self.tbl_address.setColumnWidth(2, self.tbl_address.width() * 0.1)
            self.tbl_address.setColumnWidth(3, self.tbl_address.width() * 0.2)
            self.tbl_address.setColumnWidth(4, self.tbl_address.width() * 0.2)
            self.tbl_address.setColumnWidth(5, self.tbl_address.width() * 0.2)
        else:
            self.tbl_address.setColumnWidth(0, self.tbl_address.width() * 0.2)
            self.tbl_address.setColumnWidth(1, self.tbl_address.width() * 0.15)
            self.tbl_address.setColumnWidth(2, self.tbl_address.width() * 0.15)
            self.tbl_address.setColumnWidth(3, self.tbl_address.width() * 0.15)
            self.tbl_address.setColumnWidth(4, self.tbl_address.width() * 0.2)
            self.tbl_address.setColumnWidth(5, self.tbl_address.width() * 0.15)

    @Slot()
    def btn_addRow_clicked(self):
        try:
            save_srs_list = list(srs_dict.values())

            if self.rbtn_file.isChecked():
                fileNames, types = QFileDialog.getOpenFileNames(
                    self, "选择需要转换的图形文件", os.getcwd(),
                    "图形文件(*.shp *.geojson *.dwg);;ESRI Shapefile(*.shp);;GeoJson(*.geojson);;CAD drawing(*.dwg)")
                if len(fileNames) == 0:
                    return

                for fileName in fileNames:
                    fileType = get_suffix(fileName)
                    if fileType == DataType.shapefile:
                        wks = workspaceFactory().get_factory(DataType.shapefile)
                        datasource = wks.openFromFile(fileName)
                    elif fileType == DataType.geojson:
                        wks = workspaceFactory().get_factory(DataType.geojson)
                        datasource = wks.openFromFile(fileName)
                    elif fileType == DataType.cad_dwg:
                        wks = workspaceFactory().get_factory(DataType.cad_dwg)
                        datasource = None
                    else:
                        log.error("不识别的图形文件格式!", dialog=True)
                        return None

                    if datasource is not None:
                        layer_name = wks.getLayerNames()[0]
                        in_layer = datasource.GetLayer()
                        row = self.add_layer_to_row(in_layer, fileName, layer_name)
                        # self.add_delegate_to_row(row, fileName, [layer_name], save_srs_list)
                        levelData = {
                            'layer_names': [layer_name],
                            'srs_list': save_srs_list
                        }
                        self.model.setLevelData(fileName, levelData)

                        datasource.Release()
                        datasource = None
                        in_layer = None
                    else:
                        layer_name, suffix = os.path.splitext(os.path.basename(fileName))
                        row = self.add_layer_to_row(None, fileName, layer_name)
                        levelData = {
                            'layer_names': [layer_name],
                            'srs_list': save_srs_list
                        }
                        self.model.setLevelData(fileName, levelData)
                        # self.add_delegate_to_row(row, fileName, [layer_name], save_srs_list)

            elif self.rbtn_filedb.isChecked():
                fileName = QtWidgets.QFileDialog.getExistingDirectory(self, "选择需要转换的GDB数据库",
                                                                      os.getcwd(), QFileDialog.ShowDirsOnly)
                wks = workspaceFactory().get_factory(DataType.fileGDB)
                datasource = wks.openFromFile(fileName)

                if datasource is not None:
                    lst_names = wks.getLayerNames()
                    selected_names = None
                    if len(lst_names) > 1:
                        selected_names = nameListDialog().openListDialog("请选择要转换的图层", lst_names)
                    elif len(lst_names) == 1:
                        selected_names = [lst_names[0]]

                    rows = []
                    if selected_names is not None:
                        for selected_name in selected_names:
                            layer = wks.openLayer(selected_name)
                            row = self.add_layer_to_row(layer, fileName, selected_name)
                            rows.append(row)

                        for row in rows:
                            levelData = {
                                'layer_names': lst_names,
                                'srs_list': save_srs_list
                            }
                            self.model.setLevelData(fileName, levelData)
                            # self.add_delegate_to_row(row, fileName, lst_names, save_srs_list)

            elif self.rbtn_table.isChecked():
                fileNames, types = QFileDialog.getOpenFileNames(
                    self, "选择需要转换的表格文件", os.getcwd(),
                    "表格文件(*.csv *.xlsx *.dbf);;csv文件(*.csv);;excel文件(*.xlsx);;dbf文件(*.dbf)")
                if len(fileNames) == 0:
                    return
                for fileName in fileNames:
                    fileType = get_suffix(fileName)
                    row = self.add_table_to_row(fileName)

                    if fileType == DataType.xlsx:
                        # selected_sheet = []
                        # wb = load_workbook(fileName, read_only=True)
                        # wb.close()
                        # lst_names = wb.sheetnames
                        # if len(lst_names) > 1:
                        #     selected_sheet = nameListDialog().openListDialog(
                        #         "请选择工作表(sheet)", lst_names, QAbstractItemView.SingleSelection)
                        # elif len(lst_names) == 1:
                        #     selected_sheet = lst_names[0]
                        header, bheader = read_table_header(fileName, fileType, None)
                        levelData = {
                            'is_header': bheader,
                            # 'sheet': selected_sheet[0],
                            'field_list': header,
                            'srs_list': save_srs_list
                        }
                    elif fileType == DataType.csv:
                        header, encoding, bheader = read_table_header(fileName, fileType)
                        levelData = {
                            'is_header': bheader,
                            'encoding': encoding,
                            'field_list': header,
                            'srs_list': save_srs_list
                        }
                    elif fileType == DataType.dbf:
                        header = read_table_header(fileName, fileType)
                        levelData = {
                            'is_header': True,
                            'field_list': header,
                            'srs_list': save_srs_list
                        }
                    else:
                        log.error("不识别的图形文件格式!", dialog=True)
                        return None

                    field_delegate = xyfieldDelegate(self,
                                                     [None, {'type': 'xy'}, {'type': 'xy'}, {'type': 'srs'},
                                                      {'type': 'srs'}, {'type': 'f', 'text': '请选择需要保存的文件'}])
                    self.tbl_address.setItemDelegateForRow(row, field_delegate)
                    self.model.setLevelData(fileName, levelData)
        except:
            log.error(traceback.format_exc())

    def add_table_to_row(self, fileName):
        row = self.model.rowCount(QModelIndex())
        self.model.addEmptyRow(row, 1, 0)
        in_path_index = self.tbl_address.model().index(row, self.in_path_no)
        self.tbl_address.model().setData(in_path_index, fileName)

        return row

    def add_layer_to_row(self, in_layer, fileName, layer_name):
        if in_layer is not None:
            in_srs = in_layer.GetSpatialRef()
            if in_srs is not None:
                in_srs = osr.SpatialReference(in_srs.ExportToWkt())
                srs_epsg = in_srs.GetAttrValue("AUTHORITY", 1)
                if srs_epsg is None:
                    srs_desc = None
                else:
                    srs_desc = srs_dict[int(srs_epsg)]
            else:
                srs_desc = None
        else:
            srs_desc = None

        row = self.model.rowCount(QModelIndex())
        self.model.addEmptyRow(row, 1, 0)
        in_path_index = self.tbl_address.model().index(row, self.in_path_no)
        in_layername_index = self.tbl_address.model().index(row, self.in_layername_no)
        in_srs_index = self.tbl_address.model().index(row, self.in_srs_no)

        self.tbl_address.model().setData(in_path_index, fileName)
        self.tbl_address.model().setData(in_layername_index, layer_name)
        self.tbl_address.model().setData(in_srs_index, srs_desc)

        return row

    @Slot()
    def open_addressFile(self):
        fileName, fileType = QtWidgets.QFileDialog.getOpenFileName(self, "选择坐标转换参数文件", os.getcwd(),
                                                                   "All Files(*)")
        self.txt_addressFile.setText(fileName)

        if fileName == "":
            return

        try:
            with open(fileName, 'r', encoding='utf-8') as f:
                self.paras = json.load(f)

            self.add_address_rows_from_paras()
        except:
            if self.model.rowCount(QModelIndex()) > 0:
                self.model.removeRows(self.model.rowCount(QModelIndex()) - 1, 1, QModelIndex())
            log.error("读取参数文件失败！", dialog=True)

    def add_address_rows_from_paras(self):
        imps = self.paras['exports']
        for imp in imps:
            row = self.model.rowCount(QModelIndex())
            self.model.addEmptyRow(row, 1, 0)

            if self.rbtn_table.isChecked():
                fileType = get_suffix(imp['in_path'])
                header = imp['field_list']

                if fileType == DataType.csv:
                    header, encoding, bheader = read_table_header(imp['in_path'], fileType)

                    self.model.setLevelData(imp['in_path'], {
                        'is_header': bheader,
                        'encoding': encoding,
                        'field_list': header,
                        'srs_list': imp['srs_list']
                    })
                elif fileType == DataType.xlsx:
                    header, bheader = read_table_header(imp['in_path'], fileType, None)
                    self.model.setLevelData(imp['in_path'], {
                        'is_header': bheader,
                        'field_list': header,
                        'srs_list': imp['srs_list']
                    })
                elif fileType == DataType.dbf:
                    self.model.setLevelData(imp['in_path'], {
                        'is_header': True,
                        'field_list': header,
                        'srs_list': imp['srs_list']
                    })

                x_field_index = self.tbl_address.model().index(row, 1)
                y_field_index = self.tbl_address.model().index(row, 2)
                in_srs_index = self.tbl_address.model().index(row, 3)
                out_srs_index = self.tbl_address.model().index(row, 4)
                x_field_delegate = self.tbl_address.itemDelegate(x_field_index)
                y_field_delegate = self.tbl_address.itemDelegate(y_field_index)
                in_srs_delegate = self.tbl_address.itemDelegate(in_srs_index)
                out_srs_delegate = self.tbl_address.itemDelegate(out_srs_index)

                if isinstance(x_field_delegate, xyfieldDelegate):
                    x_field_delegate.set_field_list(header)
                if isinstance(y_field_delegate, xyfieldDelegate):
                    y_field_delegate.set_field_list(header)
                if isinstance(in_srs_delegate, srsDelegate):
                    in_srs_delegate.set_srs_list(imp['srs_list'])
                if isinstance(out_srs_delegate, srsDelegate):
                    out_srs_delegate.set_srs_list(imp['srs_list'])

                field_delegate = xyfieldDelegate(self,
                                                 [None, {'type': 'xy'}, {'type': 'xy'}, {'type': 'srs'},
                                                  {'type': 'srs'}, {'type': 'f', 'text': '请选择需要保存的文件'}])
                self.tbl_address.setItemDelegateForRow(row, field_delegate)

                self.tbl_address.model().setData(self.tbl_address.model().index(row, 0), imp['in_path'])
                if imp['x_field'] in header:
                    self.tbl_address.model().setData(self.tbl_address.model().index(row, 1), imp['x_field'])
                if imp['y_field'] in header:
                    self.tbl_address.model().setData(self.tbl_address.model().index(row, 2), imp['y_field'])
                self.tbl_address.model().setData(self.tbl_address.model().index(row, 3), imp['in_srs'])
                self.tbl_address.model().setData(self.tbl_address.model().index(row, 4), imp['out_srs'])
                self.tbl_address.model().setData(self.tbl_address.model().index(row, 5), imp['out_path'])
            else:
                in_srs_index = self.tbl_address.model().index(row, self.in_srs_no)
                out_srs_index = self.tbl_address.model().index(row, self.out_srs_no)
                in_srs_delegate = self.tbl_address.itemDelegate(in_srs_index)
                out_srs_delegate = self.tbl_address.itemDelegate(out_srs_index)

                if isinstance(in_srs_delegate, srsDelegate):
                    in_srs_delegate.set_srs_list(imp['srs_list'])
                if isinstance(out_srs_delegate, srsDelegate):
                    out_srs_delegate.set_srs_list(imp['srs_list'])

                self.model.setLevelData(imp['in_path'], {
                    'layer_names': imp['layer_names'],
                    'srs_list': imp['srs_list']
                })

                self.tbl_address.model().setData(self.tbl_address.model().index(row, 0), imp['in_path'])
                self.tbl_address.model().setData(self.tbl_address.model().index(row, 1), imp['in_layer'])
                self.tbl_address.model().setData(self.tbl_address.model().index(row, 2), imp['in_srs'])
                self.tbl_address.model().setData(self.tbl_address.model().index(row, 3), imp['out_srs'])
                self.tbl_address.model().setData(self.tbl_address.model().index(row, 4), imp['out_path'])
                self.tbl_address.model().setData(self.tbl_address.model().index(row, 5), imp['out_layer'])

    @Slot(QAbstractButton)
    def buttonBox_clicked(self, button: QAbstractButton):
        if button == self.buttonBox.button(QDialogButtonBox.Ok):
            if not self.check_paras():
                return

            self.thread.start()
            self.run_process()
        elif button == self.buttonBox.button(QDialogButtonBox.Cancel):
            if self.thread.isRunning():
                self.thread.terminate()
                self.thread.wait()
            self.close()

    def check_paras(self):
        rows = range(0, self.tbl_address.model().rowCount(QModelIndex()))

        if self.rbtn_table.isChecked():
            for row in rows:
                in_path_index = self.tbl_address.model().index(row, 0, QModelIndex())
                x_field_index = self.tbl_address.model().index(row, 1, QModelIndex())
                y_field_index = self.tbl_address.model().index(row, 2, QModelIndex())
                in_srs_index = self.tbl_address.model().index(row, 3, QModelIndex())
                out_srs_index = self.tbl_address.model().index(row, 4, QModelIndex())
                out_path_index = self.tbl_address.model().index(row, 5, QModelIndex())

                in_path = str(self.tbl_address.model().data(in_path_index, Qt.DisplayRole)).strip()
                x_field = str(self.tbl_address.model().data(x_field_index, Qt.DisplayRole)).strip()
                y_field = str(self.tbl_address.model().data(y_field_index, Qt.DisplayRole)).strip()
                in_srs = str(self.tbl_address.model().data(in_srs_index, Qt.DisplayRole)).strip()
                out_srs = str(self.tbl_address.model().data(out_srs_index, Qt.DisplayRole)).strip()
                out_path = str(self.tbl_address.model().data(out_path_index, Qt.DisplayRole)).strip()

                if in_path == "":
                    log.error('第{}行缺失必要参数"输入路径"，请补全！'.format(row), dialog=True)
                    return False

                in_format = get_suffix(in_path)
                if in_format is None:
                    log.error('第{}行的输入数据格式不支持！目前只支持csv, excel和dbf'.format(row), dialog=True)
                    return False

                if x_field == "":
                    log.error('第{}行缺失必要参数"x坐标"，请补全!'.format(row))
                    return False

                if y_field == "":
                    log.error('第{}行缺失必要参数"y坐标"，请补全!'.format(row))
                    return False

                if in_srs == "":
                    log.error('第{}行缺失必要参数"输入坐标系"，请补全!'.format(row))
                    return False

                if out_srs == "":
                    log.error('第{}行缺失必要参数"输出坐标系"，请补全!'.format(row))
                    return False

                if out_path == "":
                    in_filename, in_suffix = os.path.splitext(os.path.basename(in_path))
                    out_file = self.default_outfile(in_path, in_format, in_filename, out_srs)
                    if not os.path.exists("res"):
                        os.makedirs("res")

                    out_path = os.path.join(os.path.abspath("res"), out_file)

                if is_already_opened_in_write_mode(out_path):
                    out_path = launderName(out_path)
                self.tbl_address.model().setData(out_path_index, out_path)
                log.warning('第{}行参数"输出路径"缺失数据源，自动补全为默认值{}'.format(row, out_path))
        else:
            for row in rows:
                in_path_index = self.tbl_address.model().index(row, 0, QModelIndex())
                in_layername_index = self.tbl_address.model().index(row, 1, QModelIndex())
                in_srs_index = self.tbl_address.model().index(row, 2, QModelIndex())
                out_srs_index = self.tbl_address.model().index(row, 3, QModelIndex())
                out_path_index = self.tbl_address.model().index(row, 4, QModelIndex())
                out_layername_index = self.tbl_address.model().index(row, 5, QModelIndex())

                in_path = str(self.tbl_address.model().data(in_path_index, Qt.DisplayRole)).strip()
                in_layername = str(self.tbl_address.model().data(in_layername_index, Qt.DisplayRole)).strip()
                in_srs = str(self.tbl_address.model().data(in_srs_index, Qt.DisplayRole)).strip()
                out_srs = str(self.tbl_address.model().data(out_srs_index, Qt.DisplayRole)).strip()
                out_path = str(self.tbl_address.model().data(out_path_index, Qt.DisplayRole)).strip()
                out_layername = str(self.tbl_address.model().data(out_layername_index, Qt.DisplayRole)).strip()

                if in_path == "":
                    log.error('第{}行缺失必要参数"输入路径"，请补全！'.format(row), dialog=True)
                    return False

                in_format = get_suffix(in_path)
                if in_format is None:
                    log.error('第{}行的输入数据格式不支持！目前只支持shapefile, fileGDB, geojson和cad dwg'.format(row), dialog=True)
                    return False
                else:
                    in_wks = workspaceFactory().get_factory(in_format)
                    if in_wks is None:
                        log.error("缺失图形文件引擎!")
                        return False

                    if in_wks.driver is None:
                        log.error("缺失图形文件引擎{}!".format(in_wks.driverName))
                        return False

                if in_layername == "":
                    if in_format == DataType.fileGDB:
                        log.error('第{}行参数缺失必要参数"输入图层"！')
                        return False
                    else:
                        in_layername, suffix = os.path.splitext(os.path.basename(in_path))
                        self.tbl_address.model().setData(in_layername_index, in_layername)
                        log.warning('第{}行参数缺失参数"输入图层"，已自动补全为{}'.format(row, in_layername))

                if in_srs == "":
                    log.error('第{}行缺失必要参数"输入坐标系"，请补全！'.format(row), dialog=True)
                    return False

                if out_srs == "":
                    log.error('第{}行缺失必要参数"输出坐标系"，请补全！'.format(row), dialog=True)
                    return False

                if out_layername == "":
                    out_layername = in_layername + "_" + str(out_srs)
                    self.tbl_address.model().setData(out_layername_index, out_layername)
                    log.warning('第{}行参数缺失参数"输出图层"，自动补全为默认值{}'.format(row, out_layername))

                self.autofill_outpath(row, in_path, out_path, in_layername, out_srs, in_format, out_path_index)

        return True

    def default_outfile(self, in_path, in_format, in_layername, out_srs):
        out_file = ""
        if in_format == DataType.fileGDB:
            in_filename, in_suffix = os.path.splitext(os.path.basename(in_path))
            out_file = "{}_converted.gdb".format(in_filename)
        elif in_format == DataType.geojson:
            out_file = "{}_{}_{}.geojson".format(in_layername, out_srs, encodeCurrentTime())
        elif in_format == DataType.shapefile:
            out_file = "{}_{}_{}.shp".format(in_layername, out_srs, encodeCurrentTime())
        elif in_format == DataType.cad_dwg:
            out_file = "{}_{}_{}.dwg".format(in_layername, out_srs, encodeCurrentTime())
        elif in_format == DataType.csv or in_format == DataType.xlsx or in_format == DataType.dbf:
            out_file = "{}_{}_{}.csv".format(in_layername, out_srs, encodeCurrentTime())
        return out_file

    def autofill_outpath(self, row, in_path, out_path, in_layername, out_srs, in_format, out_path_index):
        if out_path == "":
            out_file = self.default_outfile(in_path, in_format, in_layername, out_srs)

            if not os.path.exists("res"):
                os.makedirs("res")

            out_path = os.path.join(os.path.abspath("res"), out_file)
            self.tbl_address.model().setData(out_path_index, out_path)
            log.warning('第{}行参数缺失参数"输出路径"，自动补全为默认值{}'.format(row, out_path))
        else:
            out_format = get_suffix(out_path)
            if out_format is None:
                out_file = self.default_outfile(in_path, in_format, in_layername, out_srs)

                out_path = os.path.join(out_path, out_file)
                self.tbl_address.model().setData(out_path_index, out_path)
                log.warning('第{}行参数"输出路径"缺失数据源，自动补全为默认值{}'.format(row, out_path))

    def run_process(self):
        rows = range(0, self.tbl_address.model().rowCount(QModelIndex()))

        if self.rbtn_table.isChecked():
            for row in rows:
                in_path_index = self.tbl_address.model().index(row, 0, QModelIndex())
                x_field_index = self.tbl_address.model().index(row, 1, QModelIndex())
                y_field_index = self.tbl_address.model().index(row, 2, QModelIndex())
                in_srs_index = self.tbl_address.model().index(row, 3, QModelIndex())
                out_srs_index = self.tbl_address.model().index(row, 4, QModelIndex())
                out_path_index = self.tbl_address.model().index(row, 5, QModelIndex())

                in_path = str(self.tbl_address.model().data(in_path_index, Qt.DisplayRole)).strip()
                x_field = str(self.tbl_address.model().data(x_field_index, Qt.DisplayRole)).strip()
                y_field = str(self.tbl_address.model().data(y_field_index, Qt.DisplayRole)).strip()
                in_srs = str(self.tbl_address.model().data(in_srs_index, Qt.DisplayRole)).strip()
                out_srs = str(self.tbl_address.model().data(out_srs_index, Qt.DisplayRole)).strip()
                out_path = str(self.tbl_address.model().data(out_path_index, Qt.DisplayRole)).strip()

                in_srs = list(srs_dict.keys())[list(srs_dict.values()).index(in_srs)]
                out_srs = list(srs_dict.keys())[list(srs_dict.values()).index(out_srs)]
                x = self.model.levels[in_path]['field_list'].index(x_field)
                y = self.model.levels[in_path]['field_list'].index(y_field)

                # inencode = check_encoding(in_path)
                fileType = get_suffix(in_path)
                bheader = self.model.levels[in_path]['is_header']

                if fileType == DataType.csv:
                    inencode = self.model.levels[in_path]['encoding']

                    # UICore.coordTransform_table.coordTransform(in_path, inencode, bheader, x, y, in_srs, out_srs, out_path, "gbk")
                    self.coordTransformThread.transform_tbl.emit(
                        in_path, inencode, bheader, x, y, in_srs, out_srs, out_path, "gb18030")
                else:
                    self.coordTransformThread.transform_tbl.emit(
                        in_path, "gbk", bheader, x, y, in_srs, out_srs, out_path, "gb18030")
        else:
            for row in rows:
                in_path_index = self.tbl_address.model().index(row, 0, QModelIndex())
                in_layername_index = self.tbl_address.model().index(row, 1, QModelIndex())
                in_srs_index = self.tbl_address.model().index(row, 2, QModelIndex())
                out_srs_index = self.tbl_address.model().index(row, 3, QModelIndex())
                out_path_index = self.tbl_address.model().index(row, 4, QModelIndex())
                out_layername_index = self.tbl_address.model().index(row, 5, QModelIndex())

                in_path = str(self.tbl_address.model().data(in_path_index, Qt.DisplayRole)).strip()
                in_layername = str(self.tbl_address.model().data(in_layername_index, Qt.DisplayRole)).strip()
                in_srs = str(self.tbl_address.model().data(in_srs_index, Qt.DisplayRole)).strip()
                out_srs = str(self.tbl_address.model().data(out_srs_index, Qt.DisplayRole)).strip()
                out_path = str(self.tbl_address.model().data(out_path_index, Qt.DisplayRole)).strip()
                out_layername = str(self.tbl_address.model().data(out_layername_index, Qt.DisplayRole)).strip()

                in_srs = list(srs_dict.keys())[list(srs_dict.values()).index(in_srs)]
                out_srs = list(srs_dict.keys())[list(srs_dict.values()).index(out_srs)]

                self.coordTransformThread.transform.emit(in_path, in_layername, in_srs, out_path, out_layername, out_srs)

    @Slot()
    def btn_saveMetaFile_clicked(self):
        fileName, fileType = QFileDialog.getSaveFileName(self, "请选择保存的参数文件", os.getcwd(),
                                                         "json file(*.json)")

        if fileName == "":
            return

        datas = self.tbl_address.model().datas
        levels = self.tbl_address.model().levelData()
        logicRows = range(0, len(datas))

        results = []

        if self.rbtn_table.isChecked():
            for logicRow in logicRows:
                key = datas[logicRow][0]
                row_data = {
                    'in_path': datas[logicRow][0],
                    'x_field': datas[logicRow][1],
                    'y_field': datas[logicRow][2],
                    'in_srs': datas[logicRow][3],
                    'out_srs': datas[logicRow][4],
                    'out_path': datas[logicRow][5],
                    'field_list': levels[key]['field_list'],
                    'srs_list': levels[key]['srs_list']
                }
                results.append(row_data)
        else:
            for logicRow in logicRows:
                key = datas[logicRow][0]
                row_data = {
                    'in_path': datas[logicRow][0],
                    'in_layer': datas[logicRow][1],
                    'in_srs': datas[logicRow][2],
                    'out_srs': datas[logicRow][3],
                    'out_path': datas[logicRow][4],
                    'out_layer': datas[logicRow][5],
                    'layer_names': levels[key]['layer_names'],
                    'srs_list': levels[key]['srs_list']
                }
                results.append(row_data)

        res = {
            'exports': results
        }

        try:
            if fileName != '':
                with open(fileName, 'w', encoding='UTF-8') as f:
                    json.dump(res, f, ensure_ascii=False)
        except:
            log.error("文件存储路径错误，无法保存！", parent=self, dialog=True)

    @Slot()
    def rbtn_toggled(self):
        if self.rbtn_table.isChecked():
            self.table_init_table_data()
        else:
            self.table_init_spatial_data()

    def table_init(self):
        self.tbl_address.setStyle(mTableStyle())

        self.tbl_address.horizontalHeader().setStretchLastSection(True)
        self.tbl_address.verticalHeader().setDefaultSectionSize(20)
        self.tbl_address.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)  # 行高固定

        color = self.palette().color(QPalette.Button)
        self.tbl_address.horizontalHeader().setStyleSheet(
            "QHeaderView::section {{ background-color: {}}}".format(color.name()))
        self.tbl_address.verticalHeader().setStyleSheet(
            "QHeaderView::section {{ background-color: {}}}".format(color.name()))
        self.tbl_address.setStyleSheet(
            "QTableCornerButton::section {{ color: {}; border: 1px solid; border-color: {}}}".format(color.name(),
                                                                                                     color.name()))

        self.tbl_address.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.tbl_address.setEditTriggers(QAbstractItemView.SelectedClicked | QAbstractItemView.DoubleClicked)
        # self.tbl_address.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl_address.DragDropMode(QAbstractItemView.InternalMove)
        self.tbl_address.setSelectionBehavior(QAbstractItemView.SelectRows | QAbstractItemView.SelectItems)
        self.tbl_address.setDefaultDropAction(Qt.MoveAction)

        self.tbl_address.horizontalHeader().setSectionsMovable(False)
        self.tbl_address.setDragEnabled(True)
        self.tbl_address.setAcceptDrops(True)

        self.in_path_no = 0  # 输入路径字段的序号
        self.in_layername_no = 1  # 输入图层的序号
        self.out_layername_no = 5  # 输出图层的序号
        self.in_srs_no = 2
        self.out_srs_no = 3  # 输出坐标系的序号

    def table_init_spatial_data(self):
        self.model = TableModel()

        self.model.setHeaderData(0, Qt.Horizontal, "输入路径", Qt.DisplayRole)
        self.model.setHeaderData(1, Qt.Horizontal, "输入图层", Qt.DisplayRole)
        self.model.setHeaderData(2, Qt.Horizontal, "输入坐标系", Qt.DisplayRole)
        self.model.setHeaderData(3, Qt.Horizontal, "输出坐标系", Qt.DisplayRole)
        self.model.setHeaderData(4, Qt.Horizontal, "输出路径", Qt.DisplayRole)
        self.model.setHeaderData(5, Qt.Horizontal, "输出图层", Qt.DisplayRole)
        self.tbl_address.setModel(self.model)
        self.table_layout()

        layername_delegate = layernameDelegate(self, {'type': 'c'})
        self.tbl_address.setItemDelegateForColumn(1, layername_delegate)
        srs_delegate = srsDelegate(self, srs_dict.values())
        self.tbl_address.setItemDelegateForColumn(2, srs_delegate)
        self.tbl_address.setItemDelegateForColumn(3, srs_delegate)
        outputpath_delegate = outputPathDelegate(self, {'type': 'd'})
        self.tbl_address.setItemDelegateForColumn(4, outputpath_delegate)

    def table_init_table_data(self):
        self.model = TableModel()

        self.model.setHeaderData(0, Qt.Horizontal, "输入文件", Qt.DisplayRole)
        self.model.setHeaderData(1, Qt.Horizontal, "x坐标", Qt.DisplayRole)
        self.model.setHeaderData(2, Qt.Horizontal, "y坐标", Qt.DisplayRole)
        self.model.setHeaderData(3, Qt.Horizontal, "输入坐标系", Qt.DisplayRole)
        self.model.setHeaderData(4, Qt.Horizontal, "输出坐标系", Qt.DisplayRole)
        self.model.setHeaderData(5, Qt.Horizontal, "输出文件", Qt.DisplayRole)
        self.tbl_address.setModel(self.model)
        self.table_layout()

        # layername_delegate = layernameDelegate(self, {'type': 'c'})
        # self.tbl_address.setItemDelegateForColumn(1, layername_delegate)
        # srs_delegate = srsDelegate(self, srs_dict.values())
        # self.tbl_address.setItemDelegateForColumn(2, srs_delegate)
        # self.tbl_address.setItemDelegateForColumn(3, srs_delegate)
        # outputpath_delegate = outputPathDelegate(self, {'type': 'd'})
        # self.tbl_address.setItemDelegateForColumn(4, outputpath_delegate)

    @Slot()
    def btn_removeBtn_clicked(self):
        index_list = []
        selModel = self.tbl_address.selectionModel()
        if selModel is None:
            return
        if len(selModel.selectedRows()) < 1:
            return
        for model_index in selModel.selectedRows():
            index = QPersistentModelIndex(model_index)
            index_list.append(index)

        oldrow = index.row()
        for index in index_list:
            self.model.removeRows(index.row(), 1, 0)

        if self.model.rowCount(QModelIndex()) == 1:
            next_index = self.model.index(0, 0)
        else:
            next_index = self.model.index(oldrow, 0)
        # next_index = self.model.index(self.model.rowCount(QModelIndex()) - 1, 0)
        selModel.select(next_index, QItemSelectionModel.ClearAndSelect | QItemSelectionModel.Rows)
        self.tbl_address.setFocus()

    def return_url_and_layername(self, logicRow):
        in_path_index = self.tbl_address.model().index(logicRow, self.in_path_no)
        layername_index = self.tbl_address.model().index(logicRow, self.in_layername_no)
        url = self.tbl_address.model().data(in_path_index, Qt.DisplayRole)
        layername = self.tbl_address.model().data(layername_index, Qt.DisplayRole)
        return in_path_index, layername_index, url, layername

    def update_outlayername(self, index: QModelIndex, text):
        if index.column() == self.out_srs_no:
            in_layername_index = self.tbl_address.model().index(index.row(), self.in_layername_no)
            out_srs_index = self.tbl_address.model().index(index.row(), self.out_srs_no)
            out_layername_index = self.tbl_address.model().index(index.row(), self.out_layername_no)
            layer_name = self.tbl_address.model().data(in_layername_index, Qt.DisplayRole)
            # self.tbl_address.model().setData(out_srs_index, text)
            self.tbl_address.model().setData(out_layername_index, "{}_{}".format(layer_name, text))


class nameListDialog(QtWidgets.QDialog, UI.listview_dialog.Ui_Dialog):
    def __init__(self):
        super(nameListDialog, self).__init__()
        self.setupUi(self)
        self.pushButton.clicked.connect(self.pushButton_clicked)
        self.select_names = []

    def openListDialog(self, title, lst_names, selectMode=QAbstractItemView.ExtendedSelection):
        self.lv_name.setSelectionMode(selectMode)
        self.setWindowTitle(title)

        for name in lst_names:
            self.lv_name.addItem(name)

        result = self.exec_()

        if result == 1:
            return self.select_names

    def pushButton_clicked(self):
        sel_items = self.lv_name.selectedItems()
        for item in sel_items:
            self.select_names.append(item.text())
        self.done(1)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    style = QStyleFactory.create("windows")
    app.setStyle(style)
    window = Ui_Window()
    window.setWindowFlags(Qt.Window)
    window.show()
    sys.exit(app.exec_())
